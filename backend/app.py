import io
import os
import sqlite3
from datetime import date as _date
from flask import Flask, jsonify, request, render_template, send_file, abort

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

import database as db

app = Flask(__name__)

try:
    from flask_cors import CORS
    CORS(app, resources={r"/api/*": {"origins": "*"}}, supports_credentials=True)
except ImportError:
    pass

@app.before_request
def handle_options_preflight():
    if request.method == "OPTIONS":
        response = jsonify({"status": "ok"})
        origin = request.headers.get("Origin", "*")
        response.headers["Access-Control-Allow-Origin"] = origin
        response.headers["Access-Control-Allow-Credentials"] = "true"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization, X-Requested-With, X-User-ID"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
        return response, 200

@app.after_request
def add_cors_headers(response):
    origin = request.headers.get("Origin", "*")
    response.headers["Access-Control-Allow-Origin"] = origin
    response.headers["Access-Control-Allow-Credentials"] = "true"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization, X-Requested-With, X-User-ID"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
    return response

# ── Gzip compression for all responses (makes HTML/JSON ~70% smaller) ─────────
try:
    from flask_compress import Compress
    Compress(app)
except ImportError:
    pass  # Install with: pip install flask-compress

# ── Cache static files aggressively in the browser ────────────────────────────
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0  # Disabled browser cache for static files to force immediate design updates

app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "buyology-secret-session-key-185429")
app.config["SESSION_COOKIE_SAMESITE"] = "None"
app.config["SESSION_COOKIE_SECURE"] = True

import functools
from flask import session, redirect, url_for, request, jsonify, g
import json

db.init_db()


@app.route("/api/health", methods=["GET"])
def health_check():
    db_type = "PostgreSQL (Supabase)" if os.environ.get("DATABASE_URL") else "SQLite (Local)"
    db_status = "healthy"
    try:
        conn = db.get_connection()
        conn.execute("SELECT 1").fetchone()
        conn.close()
    except Exception as e:
        db_status = f"error: {str(e)}"
    return jsonify({
        "status": "ok",
        "database_type": db_type,
        "database_status": db_status,
        "environment": os.environ.get("FLASK_ENV", "production")
    })


# ── Catalog & Photo Cloud Sync API ───────────────────────────────────────────
@app.route("/api/catalog", methods=["GET"])
def api_get_catalog():
    raw_text = db.get_catalog_setting("raw_text", "")
    photos_str = db.get_catalog_setting("product_photos", "{}")
    try:
        photos = json.loads(photos_str)
    except Exception:
        photos = {}
    return jsonify({"rawText": raw_text, "productPhotos": photos})


@app.route("/api/catalog", methods=["POST"])
def api_save_catalog():
    data = request.get_json() or {}
    if "rawText" in data and isinstance(data["rawText"], str) and data["rawText"].strip():
        db.set_catalog_setting("raw_text", data["rawText"])
    if "productPhotos" in data and isinstance(data["productPhotos"], dict):
        existing_str = db.get_catalog_setting("product_photos", "{}")
        try:
            existing = json.loads(existing_str)
        except Exception:
            existing = {}
        # Safely merge incoming photos with existing DB catalog photos
        for key, photo_list in data["productPhotos"].items():
            if photo_list and isinstance(photo_list, list) and len(photo_list) > 0:
                existing[key] = photo_list
        db.set_catalog_setting("product_photos", json.dumps(existing))
    return jsonify({"ok": True, "productPhotos": json.loads(db.get_catalog_setting("product_photos", "{}"))})


@app.route("/api/photos", methods=["GET"])
def api_get_photos():
    """Dedicated endpoint to get just the product photos map (no rawText needed)."""
    photos_str = db.get_catalog_setting("product_photos", "{}")
    try:
        photos = json.loads(photos_str)
    except Exception:
        photos = {}
    return jsonify({"productPhotos": photos, "count": len(photos)})


@app.route("/api/photos", methods=["POST"])
def api_save_photos():
    """Dedicated endpoint to save/merge product photos - completely independent of rawText."""
    data = request.get_json() or {}
    incoming = data.get("productPhotos", {})
    if not isinstance(incoming, dict):
        return jsonify({"error": "productPhotos must be an object"}), 400

    existing_str = db.get_catalog_setting("product_photos", "{}")
    try:
        existing = json.loads(existing_str)
    except Exception:
        existing = {}

    updated_count = 0
    for key, photo_list in incoming.items():
        if photo_list and isinstance(photo_list, list) and len(photo_list) > 0:
            existing[key] = photo_list
            updated_count += 1

    db.set_catalog_setting("product_photos", json.dumps(existing))
    return jsonify({"ok": True, "updatedAlbums": updated_count, "totalAlbums": len(existing)})


@app.route("/api/photos/clear-all", methods=["POST"])
def api_clear_all_photos():
    """Clear all uploaded product photos from central DB catalog settings."""
    db.set_catalog_setting("product_photos", "{}")
    return jsonify({"ok": True, "message": "All product photos cleared from central database"})







@app.route("/api/admin-requests", methods=["GET"])
def api_get_admin_requests():
    return jsonify(db.get_admin_requests())


@app.route("/api/admin-requests", methods=["POST"])
def api_save_admin_request():
    data = request.get_json() or {}
    action = data.get("action")
    email = data.get("email")
    if not action or not email:
        return jsonify({"error": "action and email are required"}), 400
    db.save_admin_request(action, email)
    return jsonify({"ok": True, "data": db.get_admin_requests()})


@app.route("/api/admin-requests/reset-approved", methods=["POST"])
def api_reset_approved_list():
    """Reset the approved list to only the master email, clearing all old stale entries."""
    import json
    master = "mahinshanavas1@gmail.com"
    db.set_catalog_setting("approved_admin_emails", json.dumps([master]))
    db.set_catalog_setting("pending_admin_requests", json.dumps([]))
    return jsonify({"ok": True, "message": "Approved list reset to master only", "data": db.get_admin_requests()})


@app.route("/api/upload-photo", methods=["POST"])
def api_upload_photo():
    """
    100% Zero-Login Image Upload Endpoint.
    Uploads photo to high-speed public CDN, returns permanent public HTTPS image URL, and saves to DB.
    Zero login required for staff or master!
    """
    import urllib.request
    import time
    import json

    album_key = request.form.get("albumKey", "General")
    model_title = request.form.get("modelTitle", album_key)
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "No file provided"}), 400

    file_bytes = file.read()
    safe_name = "".join(c if c.isalnum() or c in "-_." else "_" for c in model_title)
    filename = f"{safe_name}_{int(time.time())}.jpg"

    photo_url = None

    # Step 1: Upload via TmpFiles CDN (100% Free, Zero Key Required, High-Speed Multipart)
    try:
        boundary = "----WebKitFormBoundary7MA4YWxkTrZu0gW"
        body_parts = []
        body_parts.append(f"--{boundary}\r\n".encode())
        body_parts.append(f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'.encode())
        body_parts.append(b"Content-Type: image/jpeg\r\n\r\n")
        body_parts.append(file_bytes)
        body_parts.append(f"\r\n--{boundary}--\r\n".encode())
        payload = b"".join(body_parts)

        req = urllib.request.Request(
            "https://tmpfiles.org/api/v1/upload",
            data=payload,
            headers={
                "Content-Type": f"multipart/form-data; boundary={boundary}",
                "Content-Length": str(len(payload))
            },
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            res_json = json.loads(resp.read().decode('utf-8'))
            orig_url = res_json.get("data", {}).get("url")
            if orig_url:
                # Convert tmpfiles.org/ID/filename to direct CDN image URL tmpfiles.org/dl/ID/filename
                photo_url = orig_url.replace("tmpfiles.org/", "tmpfiles.org/dl/")
    except Exception as cdn_err:
        print(f"TmpFiles upload error: {cdn_err}")
        photo_url = None

    if not photo_url:
        return jsonify({"error": "Image upload failed. Please try again."}), 500

    # Save the URL to the product_photos DB map immediately
    existing_str = db.get_catalog_setting("product_photos", "{}")
    try:
        existing = json.loads(existing_str)
    except Exception:
        existing = {}
    album = existing.get(album_key, [])
    album.append({"url": photo_url, "label": f"Photo {len(album) + 1}"})
    existing[album_key] = album
    db.set_catalog_setting("product_photos", json.dumps(existing))

    return jsonify({"ok": True, "url": photo_url, "albumKey": album_key})


@app.route("/api/photos", methods=["DELETE"])
def api_delete_photo():
    """Remove a specific photo URL from the product_photos DB map (prevents poll from restoring it)."""
    data = request.get_json() or {}
    album_key = data.get("albumKey", "")
    photo_url = data.get("url", "")
    if not album_key or not photo_url:
        return jsonify({"error": "albumKey and url are required"}), 400

    existing_str = db.get_catalog_setting("product_photos", "{}")
    try:
        existing = json.loads(existing_str)
    except Exception:
        existing = {}

    if album_key in existing:
        existing[album_key] = [p for p in existing[album_key] if p.get("url") != photo_url]
        if not existing[album_key]:
            del existing[album_key]
        db.set_catalog_setting("product_photos", json.dumps(existing))

    return jsonify({"ok": True})





# ── Helpers & Auth Decorators ─────────────────────────────────────────────────

def today_str():
    return _date.today().strftime("%d-%m-%Y")


def login_required(f):
    @functools.wraps(f)
    def decorated_function(*args, **kwargs):
        if request.method == "GET" or "user_id" in session or request.headers.get("X-User-ID"):
            return f(*args, **kwargs)
        if request.path.startswith("/api/"):
            return jsonify({"error": "Unauthorized"}), 401
        return redirect(url_for("login_view"))
    return decorated_function


def permission_required(permission_keys):
    if isinstance(permission_keys, str):
        permission_keys = [permission_keys]
    def decorator(f):
        @functools.wraps(f)
        def decorated_function(*args, **kwargs):
            if request.method == "GET" or "user_id" in session or request.headers.get("X-User-ID"):
                return f(*args, **kwargs)
            if session.get("role") == "admin":
                return f(*args, **kwargs)
            
            try:
                user_permissions = json.loads(session.get("permissions", "[]"))
            except Exception:
                user_permissions = []
                
            if not any(key in user_permissions for key in permission_keys):
                return jsonify({"error": f"Forbidden: requires one of {permission_keys} permissions"}), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def admin_required(f):
    @functools.wraps(f)
    def decorated_function(*args, **kwargs):
        if request.method == "GET" or "user_id" in session or request.headers.get("X-User-ID"):
            return f(*args, **kwargs)
        if session.get("role") != "admin":
            return jsonify({"error": "Forbidden"}), 403
        return f(*args, **kwargs)
    return decorated_function


# ── Auth & User Management Routes ─────────────────────────────────────────────

@app.route("/login", methods=["GET"])
def login_view():
    if "user_id" in session:
        return redirect(url_for("index"))
    return render_template("login.html")


@app.route("/api/auth/login", methods=["POST"])
def api_auth_login():
    data = request.get_json() or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    
    if not username or not password:
        return jsonify({"error": "Username and password are required"}), 400
        
    user = db.get_user(username)
    if not user:
        print(f"Login attempt failed: user '{username}' not found")
        return jsonify({"error": "Invalid username or password"}), 401
        
    from werkzeug.security import check_password_hash
    valid = False
    try:
        valid = check_password_hash(user.get("password_hash", ""), password)
    except Exception:
        pass

    if not valid and user.get("password_hash") == password:
        valid = True
        
    if not valid:
        print(f"Login attempt failed: invalid password for user '{username}'")
        return jsonify({"error": "Invalid username or password"}), 401
        
    # Set session
    session["user_id"] = user["id"]
    session["username"] = user["username"]
    session["role"] = user.get("role", "staff")
    session["permissions"] = user.get("permissions", "[]")
    
    return jsonify({
        "ok": True,
        "user": {
            "username": user["username"],
            "role": user.get("role", "staff"),
            "permissions": user.get("permissions", "[]")
        }
    })


@app.route("/api/auth/logout", methods=["POST"])
def api_auth_logout():
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/auth/me", methods=["GET"])
@login_required
def api_auth_me():
    return jsonify({
        "username": session.get("username"),
        "role": session.get("role"),
        "permissions": session.get("permissions")
    })


@app.route("/api/auth/register", methods=["POST"])
def api_auth_register():
    data = request.get_json() or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    role = data.get("role") or "staff"
    
    if not username or not password:
        return jsonify({"error": "Username and password are required"}), 400

    res = db.create_user(username, password, role, '["all"]' if role == 'admin' else '[]')
    if not res.get("ok"):
        return jsonify({"error": res.get("error", "Registration failed")}), 400

    user = db.get_user(username)
    session["user_id"] = user["id"]
    session["username"] = user["username"]
    session["role"] = user["role"]
    session["permissions"] = user["permissions"]

    return jsonify({
        "ok": True,
        "user": {
            "username": user["username"],
            "role": user["role"],
            "permissions": user["permissions"]
        }
    })


@app.route("/api/users", methods=["GET"])
@admin_required
def api_users_list():
    return jsonify(db.get_all_users())


@app.route("/api/users", methods=["POST"])
@admin_required
def api_users_create():
    data = request.get_json()
    if not data or not data.get("username") or not data.get("password") or not data.get("role"):
        return jsonify({"error": "Username, password, and role are required"}), 400
        
    username = data["username"].strip()
    password = data["password"]
    role = data["role"]
    permissions = json.dumps(data.get("permissions", []))
    
    res = db.create_user(username, password, role, permissions)
    if not res.get("ok"):
        return jsonify({"error": res.get("error", "Failed to create user")}), 400
        
    return jsonify({"ok": True})


@app.route("/api/users/<int:user_id>", methods=["PUT"])
@admin_required
def api_users_update(user_id):
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400
        
    username = data.get("username")
    password = data.get("password") # optional
    role = data.get("role")
    permissions = json.dumps(data.get("permissions")) if "permissions" in data else None
    
    res = db.update_user(user_id, username, password, role, permissions)
    if not res.get("ok"):
        return jsonify({"error": res.get("error", "Failed to update user")}), 400
        
    # Update current session if user updated themselves
    if session.get("user_id") == user_id:
        if username:
            session["username"] = username
        if role:
            session["role"] = role
        if permissions is not None:
            session["permissions"] = permissions
            
    return jsonify({"ok": True})


@app.route("/api/users/<int:user_id>", methods=["DELETE"])
@admin_required
def api_users_delete(user_id):
    if session.get("user_id") == user_id:
        return jsonify({"error": "Cannot delete your own account"}), 400
        
    res = db.delete_user(user_id)
    if not res.get("ok"):
        return jsonify({"error": res.get("error", "Failed to delete user")}), 400
        
    return jsonify({"ok": True})


# ── Warranty API Endpoints ───────────────────────────────────────────────────




# ── Page ─────────────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def index():
    return render_template("index.html")


# ── Combined init endpoint — ONE request loads all startup data ───────────────
@app.route("/api/init", methods=["GET"])
@login_required
def api_init():
    """Returns today's bills + stats + product list based on permissions in a single round-trip."""
    today = today_str()
    role = session.get("role")
    
    try:
        user_permissions = json.loads(session.get("permissions", "[]"))
    except Exception:
        user_permissions = []
        
    has_bills = (role == "admin" or "bills" in user_permissions)
    has_products = (role == "admin" or "products" in user_permissions)
    
    bills = db.get_bills_for_date(today) if has_bills else []
    stats = db.get_stats(today) if has_bills else {}
    products = db.get_all_products() if has_products else []
    
    return jsonify({
        "today": today,
        "bills": bills,
        "stats": stats,
        "products": products,
    })


# ── Products API ─────────────────────────────────────────────────────────────

@app.route("/api/products", methods=["GET"])
@permission_required(["products", "bills"])
def api_products_list():
    return jsonify(db.get_all_products())


@app.route("/api/products/<dta>", methods=["GET"])
@permission_required(["products", "deliveries", "bills", "display"])
def api_product_get(dta):
    # Try exact match first (case-insensitive)
    product = db.get_product(dta)
    if product:
        return jsonify(product)
        
    # If not found, try digit extraction smart match
    import re
    # Extract only the digit characters from the input search query
    digits = re.sub(r"\D", "", dta)
    if digits:
        all_products = db.get_all_products()
        # Find matches with exact digit sequence matching
        matches = [
            p for p in all_products
            if re.sub(r"\D", "", p["dta"]) == digits
        ]
        # Fallback to contains-digit-sequence matching
        if not matches:
            matches = [
                p for p in all_products
                if digits in re.sub(r"\D", "", p["dta"])
            ]
            
        if matches:
            # Return the first matching product (most relevant)
            return jsonify(matches[0])
            
    return jsonify({"error": "Not found"}), 404


@app.route("/api/products", methods=["POST"])
@permission_required("products")
def api_product_create():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No payload"}), 400
    if isinstance(data, list):
        for item in data:
            if item.get("dta"):
                db.upsert_product(
                    dta=item["dta"],
                    brand=item.get("brand", ""),
                    model=item.get("model", ""),
                    price=item.get("price", 0),
                    image_url=item.get("image_url", item.get("photo", ""))
                )
        return jsonify({"ok": True, "count": len(data)})
    if not data.get("dta"):
        return jsonify({"error": "dta is required"}), 400
    db.upsert_product(
        dta=data["dta"],
        brand=data.get("brand", ""),
        model=data.get("model", ""),
        price=data.get("price", 0),
        image_url=data.get("image_url", data.get("photo", ""))
    )
    return jsonify({"ok": True, "dta": data["dta"].upper()})


@app.route("/api/products/<dta>", methods=["PUT"])
@permission_required("products")
def api_product_update(dta):
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data"}), 400
    db.upsert_product(
        dta=dta,
        brand=data.get("brand", ""),
        model=data.get("model", ""),
        price=data.get("price", 0),
        image_url=data.get("image_url", data.get("photo", ""))
    )
    return jsonify({"ok": True})


@app.route("/api/products/<dta>", methods=["DELETE"])
@permission_required("products")
def api_product_delete(dta):
    db.delete_product(dta)
    return jsonify({"ok": True})

@app.route("/api/products/search", methods=["GET"])
@permission_required("products")
def api_product_search():
    """Return all products whose brand matches AND model prefix matches,
    excluding the current DTA (so you can detect same-product variants).
    Query: ?brand=HP&model=Pavilion+15&exclude_dta=DTA9999
    """
    brand       = (request.args.get("brand") or "").strip().lower()
    model       = (request.args.get("model") or "").strip().lower()
    exclude_dta = (request.args.get("exclude_dta") or "").strip().upper()
    if not brand or not model:
        return jsonify([])
    all_products = db.get_all_products()
    matches = [
        p for p in all_products
        if p["brand"].lower() == brand
        and (
            p["model"].lower().startswith(model[:10])
            or model.startswith(p["model"].lower()[:10])
        )
        and p["dta"].upper() != exclude_dta
    ]
    return jsonify(matches)


def parse_brand_and_model(item_name):
    item_name = item_name.strip()
    if not item_name:
        return "", ""
    
    # Split by whitespace to find the first word (brand)
    parts = item_name.split(None, 1)
    brand = parts[0]
    
    # Keep the brand name in the model field also
    model = item_name
    
    return brand, model


@app.route("/api/products/upload-preview", methods=["POST"])
@permission_required("products")
def api_products_upload_preview():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    
    file = request.files["file"]
    filename = file.filename
    if not filename:
        return jsonify({"error": "No file selected"}), 400
    
    ext = os.path.splitext(filename)[1].lower()
    parsed_products = []
    
    try:
        if ext == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            
            # Read rows
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(row)
                
        elif ext == ".csv":
            import csv
            # Read CSV
            stream = io.StringIO(file.stream.read().decode("utf-8", errors="ignore"), newline=None)
            csv_reader = csv.reader(stream)
            rows = list(csv_reader)
            
        else:
            return jsonify({"error": "Unsupported file format. Please upload .xlsx or .csv"}), 400
            
        if not rows:
            return jsonify({"error": "The uploaded file is empty"}), 400
            
        # Find headers
        header_row = None
        header_idx = -1
        # Try to find a row containing 'item' and 'name' or similar (using normalized exact matching)
        for idx, row in enumerate(rows[:5]):
            if not row:
                continue
            row_normalized = [str(cell).lower().replace("_", " ").replace("-", " ").strip() if cell is not None else "" for cell in row]
            # Check if any cell matches 'item' or 'dta'
            has_item = any(cell in ["item", "dta", "code", "dta code", "item code"] for cell in row_normalized)
            has_name = any(cell in ["item name", "product name", "model", "description", "product", "name"] for cell in row_normalized)
            if has_item and has_name:
                header_row = row_normalized
                header_idx = idx
                break
                
        # If no clear headers found, assume row 0 is header
        if header_idx == -1:
            header_row = [str(cell).lower().replace("_", " ").replace("-", " ").strip() if cell is not None else "" for cell in rows[0]]
            header_idx = 0
            
        # Map columns
        dta_col_idx = -1
        name_col_idx = -1
        price_col_idx = -1
        brand_col_idx = -1
        
        for idx, val in enumerate(header_row):
            if val in ["dta code", "item code", "dta", "item"]:
                dta_col_idx = idx
            elif val in ["item name", "product name", "model", "description", "product", "name"]:
                name_col_idx = idx
            elif val in ["price", "selling price", "rate", "cost", "value", "unit price"]:
                price_col_idx = idx
            elif val in ["brand", "manufacturer", "make"]:
                brand_col_idx = idx
                
        # Fallbacks if perfect match not found
        if dta_col_idx == -1:
            dta_col_idx = 0
        if name_col_idx == -1:
            name_col_idx = min(1, len(header_row) - 1)
            
        # Parse data rows (start after header_idx)
        for row in rows[header_idx + 1:]:
            if not row or len(row) <= max(dta_col_idx, name_col_idx):
                continue
                
            dta_val = row[dta_col_idx]
            name_val = row[name_col_idx]
            
            if dta_val is None or name_val is None:
                continue
                
            dta = str(dta_val).strip().upper()
            item_name = str(name_val).strip()
            
            if not dta or not item_name:
                continue
                
            # If DTA value is header itself (just in case), skip
            if dta.lower() in ["item", "dta", "code", "dta code", "item name", "product name"]:
                continue
                
            # Parse price if column exists
            price = 0.0
            if price_col_idx != -1 and price_col_idx < len(row):
                p_val = row[price_col_idx]
                if p_val is not None:
                    try:
                        price = float(p_val)
                    except ValueError:
                        pass
                        
            # Parse brand and model
            if brand_col_idx != -1 and brand_col_idx < len(row) and row[brand_col_idx] is not None:
                brand = str(row[brand_col_idx]).strip()
                model = item_name
            else:
                brand, model = parse_brand_and_model(item_name)
                
            parsed_products.append({
                "dta": dta,
                "brand": brand,
                "model": model,
                "price": price
            })
            
        return jsonify({
            "ok": True,
            "filename": filename,
            "total_rows": len(parsed_products),
            "preview": parsed_products[:5],
            "products": parsed_products
        })
        
    except Exception as e:
        return jsonify({"error": f"Error parsing file: {str(e)}"}), 500


@app.route("/api/products/upload-save", methods=["POST"])
@permission_required("products")
def api_products_upload_save():
    data = request.get_json()
    if not data or "products" not in data:
        return jsonify({"error": "No products list provided"}), 400
        
    products = data["products"]
    count = 0
    for p in products:
        dta = (p.get("dta") or "").strip().upper()
        brand = (p.get("brand") or "").strip()
        model = (p.get("model") or "").strip()
        price = float(p.get("price") or 0)
        
        if dta and brand and model:
            db.upsert_product(dta, brand, model, price)
            count += 1
            
    return jsonify({"ok": True, "count": count})




# ── Bills API ─────────────────────────────────────────────────────────────────

@app.route("/api/bills", methods=["GET"])
@permission_required(["bills", "history"])
def api_bills_list():
    all_bills = request.args.get("all", "false").lower() == "true"
    if all_bills:
        conn = db.get_connection()
        rows = conn.execute(
            "SELECT * FROM bills ORDER BY id ASC"
        ).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    date = request.args.get("date", today_str())
    return jsonify(db.get_bills_for_date(date))


@app.route("/api/bills/dates", methods=["GET"])
@permission_required(["bills", "history", "dashboard"])
def api_bill_dates():
    return jsonify(db.get_all_bill_dates())


@app.route("/api/bills", methods=["POST"])
@permission_required("bills")
def api_bill_create():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data"}), 400

    import json
    # Auto-save all products in the list to DB if provided
    products = data.get("products", [])
    if products:
        for p in products:
            p_dta = (p.get("dta") or "").strip().upper()
            if p_dta and p.get("brand") and p.get("model"):
                db.upsert_product(p_dta, p["brand"], p["model"], float(p.get("price") or 0.0))
        data["products_json"] = json.dumps(products)
    else:
        # Fallback to single product auto-save
        dta = (data.get("dta") or "").strip().upper()
        if dta and data.get("brand") and data.get("model"):
            db.upsert_product(dta, data["brand"], data["model"], data.get("price", 0))

    # For exchange: also save new product DTA
    exch_dta = (data.get("exch_new_dta") or "").strip().upper()
    if exch_dta and data.get("exch_new_brand") and data.get("exch_new_model"):
        db.upsert_product(
            exch_dta,
            data["exch_new_brand"],
            data["exch_new_model"],
            data.get("exch_new_price", 0),
        )

    if not data.get("date"):
        data["date"] = today_str()

    new_id = db.create_bill(data)

    warranty_claim_id = data.get("warranty_claim_id")
    if warranty_claim_id:
        conn = db.get_connection()
        tx_type = data.get("transaction_type", "Sale")
        if tx_type == "Exchange":
            conn.execute(
                """UPDATE warranty_claims 
                   SET status = 'Exchanged',
                       exch_new_dta = ?,
                       exch_new_brand = ?,
                       exch_new_model = ?,
                       exch_balance = ?,
                       action_date = ?
                   WHERE id = ?""",
                (
                    (data.get("exch_new_dta") or "").upper(),
                    data.get("exch_new_brand", ""),
                    data.get("exch_new_model", ""),
                    float(data.get("exch_balance") or 0.0),
                    data.get("date") or today_str(),
                    int(warranty_claim_id)
                )
            )
            conn.commit()
        elif tx_type == "Return":
            conn.execute(
                """UPDATE warranty_claims 
                   SET status = 'Returned',
                       refund_amount = ?,
                       action_date = ?
                   WHERE id = ?""",
                (
                    float(data.get("price") or 0.0),
                    data.get("date") or today_str(),
                    int(warranty_claim_id)
                )
            )
            conn.commit()
        conn.close()

    return jsonify({"ok": True, "id": new_id})


@app.route("/api/bills/<int:bill_id>", methods=["PUT"])
@permission_required(["bills", "history"])
def api_bill_update(bill_id):
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data"}), 400

    import json
    # Auto-save all products in the list to DB if provided
    products = data.get("products", [])
    if products:
        for p in products:
            p_dta = (p.get("dta") or "").strip().upper()
            if p_dta and p.get("brand") and p.get("model"):
                db.upsert_product(p_dta, p["brand"], p["model"], float(p.get("price") or 0.0))
        data["products_json"] = json.dumps(products)
    else:
        # Fallback to single product auto-save
        dta = (data.get("dta") or "").strip().upper()
        if dta and data.get("brand") and data.get("model"):
            db.upsert_product(dta, data["brand"], data["model"], data.get("price", 0))

    # For exchange: also save new product DTA
    exch_dta = (data.get("exch_new_dta") or "").strip().upper()
    if exch_dta and data.get("exch_new_brand") and data.get("exch_new_model"):
        db.upsert_product(
            exch_dta,
            data["exch_new_brand"],
            data["exch_new_model"],
            data.get("exch_new_price", 0),
        )

    db.update_bill(bill_id, data)

    return jsonify({"ok": True})


@app.route("/api/bills/<int:bill_id>", methods=["DELETE"])
@permission_required(["bills", "history"])
def api_bill_delete(bill_id):
    db.delete_bill(bill_id)
    return jsonify({"ok": True})


@app.route("/api/bills/lookup/<int:bill_id>", methods=["GET"])
@permission_required(["bills", "history", "deliveries"])
def api_bill_lookup(bill_id):
    bill = db.get_bill(bill_id)
    if not bill:
        return jsonify({"error": "Bill not found"}), 404
    return jsonify(bill)


# ── Stats API ─────────────────────────────────────────────────────────────────

@app.route("/api/stats", methods=["GET"])
@permission_required("dashboard")
def api_stats():
    date = request.args.get("date")
    return jsonify(db.get_stats(date))


@app.route("/api/stats/monthly", methods=["GET"])
@permission_required("dashboard")
def api_stats_monthly():
    month = request.args.get("month")
    if not month:
        from datetime import date as _date
        month = _date.today().strftime("%m-%Y")
    return jsonify(db.get_monthly_stats(month))


@app.route("/api/bills/monthly", methods=["GET"])
@permission_required("dashboard")
def api_bills_monthly():
    month = request.args.get("month")
    if not month:
        from datetime import date as _date
        month = _date.today().strftime("%m-%Y")
    conn = db.get_connection()
    rows = conn.execute(
        "SELECT * FROM bills WHERE date LIKE ? ORDER BY date ASC, id ASC",
        (f"%-{month}",)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


# ── Export API ────────────────────────────────────────────────────────────────

@app.route("/api/export", methods=["GET"])
@permission_required("history")
def api_export():
    date = request.args.get("date", today_str())
    bills = db.get_bills_for_date(date)

    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side
        )
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Sales {date}"

    # ── Colours ───────────────────────────────────────────────────────────────
    yellow_fill   = PatternFill("solid", fgColor="FFD700")
    blue_fill     = PatternFill("solid", fgColor="1E3A5F")
    alt_fill      = PatternFill("solid", fgColor="EBF2FF")
    exch_fill     = PatternFill("solid", fgColor="DDEEFF")   # light blue for exchange rows
    collect_fill  = PatternFill("solid", fgColor="D4EDDA")   # green for positive balance
    refund_fill   = PatternFill("solid", fgColor="F8D7DA")   # red for negative balance

    white_font    = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    header_font   = Font(bold=True, size=13, name="Calibri")
    bold_font     = Font(bold=True, name="Calibri")
    green_font    = Font(bold=True, color="155724", name="Calibri")
    red_font      = Font(bold=True, color="721C24", name="Calibri")
    thin_border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    COLS = [
        "#", "Customer", "Brand", "Model", "DTA", "Price (AED)",
        "MOP", "Note", "Type", "Platform", "Delivery", "Mixed Detail",
        # Exchange columns
        "Exch — New Product", "Exch — New DTA", "Exch — New Price (AED)",
        "Exch — Old Product", "Exch — Old DTA", "Exch — Old Price (AED)",
        "Exch — Balance (AED)",
    ]
    NUM_COLS = len(COLS)

    # ── Row 1: Date header ────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    cell = ws.cell(row=1, column=1, value=f"Sales Sheet — {date}")
    cell.fill = yellow_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Row 2: Column headers ─────────────────────────────────────────────────
    for ci, col in enumerate(COLS, 1):
        cell = ws.cell(row=2, column=ci, value=col)
        cell.fill = blue_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[2].height = 30

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, bill in enumerate(bills, 1):
        row_num = ri + 2
        is_exchange = bill["transaction_type"] == "Exchange"
        row_fill = exch_fill if is_exchange else (alt_fill if ri % 2 == 0 else None)

        mixed_detail = ""
        if bill["payment_mode"] == "Mixed":
            parts = []
            if bill["mixed_cash"]:   parts.append(f"Cash: AED {bill['mixed_cash']:.0f}")
            if bill["mixed_card"]:   parts.append(f"Card: AED {bill['mixed_card']:.0f}")
            if bill["mixed_tabby"]:  parts.append(f"Tabby: AED {bill['mixed_tabby']:.0f}")
            if bill["mixed_tamara"]: parts.append(f"Tamara: AED {bill['mixed_tamara']:.0f}")
            mixed_detail = " | ".join(parts)

        # For exchange, show new product in main brand/model/dta/price columns
        import json
        p_list = []
        try:
            p_list = json.loads(bill.get("products_json") or "[]")
        except Exception:
            pass

        if p_list and not is_exchange:
            brand = " + ".join(filter(None, [p.get("brand") for p in p_list]))
            
            model_parts = []
            for p in p_list:
                p_model = p.get("model") or ""
                p_qty = int(p.get("quantity") or 1)
                if p_qty > 1:
                    model_parts.append(f"{p_model} (x{p_qty})")
                else:
                    model_parts.append(p_model)
            model = " + ".join(filter(None, model_parts))
            
            dta_parts = []
            for p in p_list:
                p_dta = p.get("dta") or ""
                p_qty = int(p.get("quantity") or 1)
                if p_qty > 1:
                    dta_parts.append(f"{p_dta} (x{p_qty})")
                else:
                    dta_parts.append(p_dta)
            dta = " + ".join(filter(None, dta_parts))
            
            price = sum(float(p.get("price") or 0.0) * int(p.get("quantity") or 1) for p in p_list)
            source = " + ".join(filter(None, [p.get("source") or "Inventory" for p in p_list]))
        else:
            brand = bill["exch_new_brand"] or bill["brand"] if is_exchange else bill["brand"]
            model = bill["exch_new_model"] or bill["model"] if is_exchange else bill["model"]
            dta   = bill["exch_new_dta"]   or bill["dta"]   if is_exchange else bill["dta"]
            price = bill["exch_new_price"] or bill["price"] if is_exchange else bill["price"]
            source = bill["source"] or "Inventory"

        # Exchange detail columns
        exch_new_product = f"{bill['exch_new_brand']} {bill['exch_new_model']}".strip() if is_exchange else ""
        exch_new_dta     = bill["exch_new_dta"] if is_exchange else ""
        exch_new_price   = bill["exch_new_price"] if is_exchange else ""
        exch_old_product = f"{bill['exch_old_brand']} {bill['exch_old_model']}".strip() if is_exchange else ""
        exch_old_dta     = bill["exch_old_dta"] if is_exchange else ""
        exch_old_price   = bill["exch_old_price"] if is_exchange else ""
        exch_old_source  = bill["exch_old_source"] if is_exchange else ""
        exch_balance     = bill["exch_balance"] if is_exchange else ""

        row_data = [
            ri,
            bill["customer_name"],
            brand,
            model,
            dta,
            price,
            bill["payment_mode"],
            bill["note"],
            bill["transaction_type"],
            bill["platform"],
            "Yes" if bill["delivery"] else "No",
            mixed_detail,
            exch_new_product,
            exch_new_dta,
            exch_new_price,
            exch_old_product,
            exch_old_dta,
            exch_old_price,
            exch_balance,
        ]

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if row_fill:
                cell.fill = row_fill

        # Color-code the balance cell (column 19)
        if is_exchange and exch_balance != "":
            bal_cell = ws.cell(row=row_num, column=19)
            if isinstance(exch_balance, (int, float)):
                if exch_balance > 0:
                    bal_cell.fill   = collect_fill
                    bal_cell.font   = green_font
                    bal_cell.value  = f"+{exch_balance:.2f}  ↑ Collect"
                elif exch_balance < 0:
                    bal_cell.fill   = refund_fill
                    bal_cell.font   = red_font
                    bal_cell.value  = f"{exch_balance:.2f}  ↓ Refund"
                else:
                    bal_cell.font   = bold_font
                    bal_cell.value  = "0.00  ✓ Even"

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [4, 18, 14, 18, 12, 14, 10, 26, 10, 18, 9, 28,
              22, 12, 16, 22, 12, 16, 18]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    # Freeze top 2 rows
    ws.freeze_panes = "A3"

    # ── Save to buffer ────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Sales_{date}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ── PDF Export ────────────────────────────────────────────────────────────────

@app.route("/api/export-pdf", methods=["GET"])
@permission_required("history")
def api_export_pdf():
    date = request.args.get("date", today_str())
    bills = db.get_bills_for_date(date)

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable, Image
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    except ImportError:
        return jsonify({"error": "reportlab not installed. Run: py -m pip install reportlab"}), 500

    buf = io.BytesIO()

    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    # ── Colour palette ─────────────────────────────────────────────────────────
    BRAND_COLOR  = colors.HexColor("#402F75")
    DARK         = colors.HexColor("#1F1B2C")
    LIGHT_GRAY   = colors.HexColor("#f5f7fa")
    MID_GRAY     = colors.HexColor("#dee2e6")
    GREEN        = colors.HexColor("#28a745")
    GREEN_BG     = colors.HexColor("#e8f5e9")
    RED_C        = colors.HexColor("#dc3545")
    RED_BG       = colors.HexColor("#fff0f0")
    BLUE         = colors.HexColor("#0d6efd")
    EXCH_BG      = colors.HexColor("#ECE9F4")
    ALT_ROW      = colors.HexColor("#F8F7FA")
    WHITE        = colors.white
    MUTED        = colors.HexColor("#6c757d")
    ORANGE       = colors.HexColor("#FCBC12")

    # ── Text styles ────────────────────────────────────────────────────────────
    title_style = ParagraphStyle(
        "title", fontName="Helvetica-Bold", fontSize=20,
        textColor=WHITE, alignment=TA_LEFT, leading=24,
    )
    sub_style = ParagraphStyle(
        "sub", fontName="Helvetica-Bold", fontSize=9.5,
        textColor=colors.HexColor("#FCBC12"), alignment=TA_LEFT,
    )
    sub_right_style = ParagraphStyle(
        "sub_right", parent=sub_style, alignment=TA_RIGHT,
    )
    section_style = ParagraphStyle(
        "section", fontName="Helvetica-Bold", fontSize=11,
        textColor=DARK, spaceBefore=6, spaceAfter=4,
    )
    note_style = ParagraphStyle(
        "note", fontName="Helvetica-Oblique", fontSize=8,
        textColor=MUTED,
    )
    th_style = ParagraphStyle(
        "th", fontName="Helvetica-Bold", fontSize=8,
        textColor=WHITE, alignment=TA_CENTER,
    )
    td  = ParagraphStyle("td",  fontName="Helvetica", fontSize=8, textColor=DARK, alignment=TA_LEFT)
    tdc = ParagraphStyle("tdc", fontName="Helvetica", fontSize=8, textColor=DARK, alignment=TA_CENTER)
    tdr = ParagraphStyle("tdr", fontName="Helvetica-Bold", fontSize=8, textColor=DARK, alignment=TA_RIGHT)
    td_muted  = ParagraphStyle("tdm", fontName="Helvetica", fontSize=7.5, textColor=MUTED, alignment=TA_LEFT)
    td_green  = ParagraphStyle("tdg", fontName="Helvetica-Bold", fontSize=9, textColor=GREEN, alignment=TA_RIGHT)
    td_orange = ParagraphStyle("tdo", fontName="Helvetica-Bold", fontSize=9, textColor=ORANGE, alignment=TA_RIGHT)
    stat_label = ParagraphStyle("statlbl", fontName="Helvetica", fontSize=7, textColor=MUTED, alignment=TA_RIGHT)

    # ── Helpers ────────────────────────────────────────────────────────────────
    def aed(v):
        try:
            return f"AED {float(v):,.2f}"
        except Exception:
            return "AED 0.00"

    def get_model_name(full_model):
        if not full_model:
            return "—"
        return full_model.split("|")[0].strip()

    def mop_text(b):
        if b["payment_mode"] != "Mixed":
            return b["payment_mode"]
        parts = []
        if b["mixed_cash"]:   parts.append(f"Cash {b['mixed_cash']:.0f}")
        if b["mixed_card"]:   parts.append(f"Card {b['mixed_card']:.0f}")
        if b["mixed_tabby"]:  parts.append(f"Tabby {b['mixed_tabby']:.0f}")
        if b["mixed_tamara"]: parts.append(f"Tamara {b['mixed_tamara']:.0f}")
        return " + ".join(parts) if parts else "Mixed"

    # ── Summary numbers ────────────────────────────────────────────────────────
    sales     = [b for b in bills if b["transaction_type"] == "Sale"]
    returns   = [b for b in bills if b["transaction_type"] == "Return"]
    exchanges = [b for b in bills if b["transaction_type"] == "Exchange"]
    jenny_bills = [b for b in bills if b.get("jenny") == 1]
    
    # Gross sales (before returns, excluding Jenny)
    gross_rev = 0.0
    for b in bills:
        if b.get("jenny") == 1:
            continue
        tx = b["transaction_type"]
        if tx == "Sale":
            gross_rev += b["price"] or 0.0
        elif tx == "Exchange":
            bal = b["exch_balance"] or 0.0
            if bal > 0:
                gross_rev += bal

    # Total return value
    total_return_val = sum(b["price"] or 0.0 for b in returns if not b.get("jenny"))

    # Total Revenue: Sales + positive Exchange balances - Returns (excluding Jenny)
    total_rev = gross_rev - total_return_val

    # In-Store, Delivery & Other Revenue: Sales + positive Exchange balances - Returns (excluding Jenny)
    instore_rev = 0.0
    delivery_rev = 0.0
    instore_count = 0
    delivery_count = 0
    other_rev = 0.0
    for b in bills:
        if b.get("jenny") == 1:
            continue
        tx = b["transaction_type"]
        val = b["price"] or 0.0 if tx in ["Sale", "Return"] else (b["exch_balance"] or 0.0)
        if tx == "Exchange" and val < 0:
            val = 0.0
        
        sign = -1 if tx == "Return" else 1
        
        if b["delivery"] == 1 or b["delivery"] is True:
            delivery_rev += sign * val
            if tx != "Return":
                delivery_count += 1
        else:
            mop = b["payment_mode"]
            is_other = False
            if tx == "Sale" and mop in ["Unknown", "Nil"]:
                is_other = True
            
            if is_other:
                other_rev += sign * val
            else:
                instore_rev += sign * val
                if tx != "Return":
                    instore_count += 1

    # MOP: Sales use full price; Exchanges use balance paid by customer (positive only); Returns subtract (excluding Jenny)
    mop_cash = mop_card = mop_tabby = mop_tamara = mop_bank = 0.0
    for b in bills:
        if b.get("jenny") == 1:
            continue
        tx = b["transaction_type"]
        sign = -1 if tx == "Return" else 1

        if tx in ["Sale", "Return"]:
            amount = b["price"] or 0.0
        else:  # Exchange
            bal = b["exch_balance"] or 0.0
            amount = bal if bal > 0 else 0.0  # positive = customer paid extra

        mop = b["payment_mode"]
        if mop == "Mixed":
            mop_cash   += sign * (b["mixed_cash"]   or 0.0)
            mop_card   += sign * (b["mixed_card"]   or 0.0)
            mop_tabby  += sign * (b["mixed_tabby"]  or 0.0)
            mop_tamara += sign * (b["mixed_tamara"] or 0.0)
            mop_bank   += sign * (b["mixed_bank"]   or 0.0)
        elif mop == "Cash":
            mop_cash   += sign * amount
        elif mop == "Card":
            mop_card   += sign * amount
        elif mop == "Tabby":
            mop_tabby  += sign * amount
        elif mop == "Tamara":
            mop_tamara += sign * amount
        elif mop in ["Bank Transfer", "Bank"]:
            mop_bank   += sign * amount

    story = []

    # ══════════════════════════════════════════════════════════════════════════
    # BRANDED HEADER BANNER
    # ══════════════════════════════════════════════════════════════════════════
    logo_path = os.path.join(os.path.dirname(__file__), "static", "logo.png")
    logo_widget = None
    if os.path.exists(logo_path):
        try:
            logo_widget = Image(logo_path, width=42*mm, height=9*mm)
        except Exception:
            pass

    header_left = logo_widget if logo_widget else Paragraph("<b>BUYOLOGY</b>", title_style)

    header_data = [[
        header_left,
        Paragraph(f"Daily Sales Report — {date}", sub_style),
        Paragraph(f"Generated: {_date.today().strftime('%d %B %Y')}", sub_right_style),
    ]]
    header_table = Table(header_data, colWidths=[80*mm, 120*mm, 67*mm])
    header_table.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, -1), BRAND_COLOR),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING",   (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
        ("ALIGN",        (2, 0), (2, 0),   "RIGHT"),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 5 * mm))

    # ══════════════════════════════════════════════════════════════════════════
    # SUMMARY STATS BOX
    # ══════════════════════════════════════════════════════════════════════════
    stat_bold  = ParagraphStyle("statb", fontName="Helvetica-Bold", fontSize=11, alignment=TA_CENTER)
    stat_lbl2  = ParagraphStyle("statl", fontName="Helvetica", fontSize=8, textColor=MUTED, alignment=TA_CENTER)
    stat_red   = ParagraphStyle("statr", fontName="Helvetica-Bold", fontSize=11, textColor=RED_C, alignment=TA_CENTER)
    stat_grn   = ParagraphStyle("statg", fontName="Helvetica-Bold", fontSize=11, textColor=GREEN, alignment=TA_CENTER)
    stat_mut   = ParagraphStyle("statm", fontName="Helvetica-Bold", fontSize=11, textColor=MUTED, alignment=TA_CENTER)

    def stat_cell(value, label, style=None):
        s = style if style else stat_bold
        return [Paragraph(str(value), s), Paragraph(label, stat_lbl2)]

    # Row 1: Revenue overview
    row1 = [
        stat_cell(f"{len(sales) + len(exchanges)}", f"Total Sales\n(incl. {len(exchanges)} exchanges)"),
        stat_cell(aed(total_rev),   "Net Revenue\n(after returns)",  stat_grn),
        stat_cell(aed(gross_rev),   "Gross Sales\n(before returns)"),
        stat_cell(
            f"{len(returns)}  −{aed(total_return_val)}",
            "Returns\n(deducted from total)",
            stat_red
        ),
        stat_cell(
            f"{len(jenny_bills)}",
            "Jenny Txns\n(not in revenue)",
            stat_mut
        ),
    ]
    # Row 2: Channel & MOP
    row2 = [
        stat_cell(f"{aed(instore_rev)}\n{instore_count} orders", "In-Store Revenue"),
        stat_cell(f"{aed(delivery_rev)}\n{delivery_count} orders", "Delivery Revenue"),
        stat_cell(aed(mop_cash),   "Cash"),
        stat_cell(aed(mop_card),   "Card"),
        stat_cell(aed(mop_tabby),  "Tabby"),
        stat_cell(aed(mop_tamara), "Tamara"),
        stat_cell(aed(mop_bank),   "Bank Transfer"),
    ]

    sum_tbl1 = Table([row1], colWidths=["20%"] * 5)
    sum_tbl1.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), LIGHT_GRAY),
        ("BOX",           (0, 0), (-1, -1), 0.5, MID_GRAY),
        ("INNERGRID",     (0, 0), (-1, -1), 0.5, MID_GRAY),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("BACKGROUND",    (1, 0), (2, 0),   GREEN_BG),
        ("BACKGROUND",    (3, 0), (3, 0),   RED_BG),
        ("BACKGROUND",    (4, 0), (4, 0),   LIGHT_GRAY),
    ]))
    sum_tbl2 = Table([row2], colWidths=[
        "15%", "15%", "14.3%", "14.3%", "14.3%", "14.3%", "12.8%"
    ])
    sum_tbl2.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), LIGHT_GRAY),
        ("BOX",           (0, 0), (-1, -1), 0.5, MID_GRAY),
        ("INNERGRID",     (0, 0), (-1, -1), 0.5, MID_GRAY),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(sum_tbl1)
    story.append(Spacer(1, 2 * mm))
    story.append(sum_tbl2)
    story.append(Spacer(1, 6 * mm))

    # ══════════════════════════════════════════════════════════════════════════
    # TRANSACTIONS TABLE
    # ══════════════════════════════════════════════════════════════════════════
    story.append(Paragraph("Transaction Details", section_style))
    story.append(HRFlowable(width="100%", thickness=1, color=BRAND_COLOR, spaceAfter=4))

    HEADERS    = ["#", "Customer", "Product / Exchange Details", "DTA", "Price / Balance", "MOP", "Type", "Del.", "Note"]
    COL_WIDTHS = [8*mm, 30*mm, 87*mm, 22*mm, 30*mm, 30*mm, 18*mm, 12*mm, 30*mm]

    table_data = [[Paragraph(h, th_style) for h in HEADERS]]

    for i, b in enumerate(bills):
        is_exch = b["transaction_type"] == "Exchange"
        tx_type = b["transaction_type"]

        if is_exch:
            # ── Exchange: two clearly labelled blocks ──────────────────────
            new_brand = (b["exch_new_brand"] or "").strip()
            new_model = (b["exch_new_model"] or "").strip()
            new_dta   = (b["exch_new_dta"]   or b["dta"] or "—").strip()
            new_price = b["exch_new_price"] or b["price"] or 0
            old_brand = (b["exch_old_brand"] or "").strip()
            old_model = (b["exch_old_model"] or "").strip()
            old_dta   = (b["exch_old_dta"]   or "—").strip()
            old_price = b["exch_old_price"] or 0
            bal       = b["exch_balance"] or 0

            new_line = f"{new_brand} {get_model_name(new_model)}".strip() or "—"
            old_line = f"{old_brand} {get_model_name(old_model)}".strip() or "—"

            details_text = (
                f"<b><font color='#0d6efd'>NEW (Sold):</font></b> {new_line}<br/>"
                f"<font size='7' color='#6c757d'>&nbsp;&nbsp;DTA: {new_dta} &nbsp;|&nbsp; Selling Price: {aed(new_price)}</font><br/>"
                f"<b><font color='#714b67'>OLD (Traded In):</font></b> {old_line}<br/>"
                f"<font size='7' color='#6c757d'>&nbsp;&nbsp;DTA: {old_dta} &nbsp;|&nbsp; Accepted Value: {aed(old_price)}</font>"
            )
            product_cell = Paragraph(details_text, td)
            dta_cell = Paragraph(new_dta, tdc)

            # Balance cell — what customer actually paid / received
            if bal > 0:
                bal_cell = [
                    Paragraph(aed(bal), td_green),
                    Paragraph("Customer Pays", stat_label),
                ]
            elif bal < 0:
                bal_cell = [
                    Paragraph(aed(abs(bal)), td_orange),
                    Paragraph("Refund to Customer", stat_label),
                ]
            else:
                bal_cell = [
                    Paragraph("AED 0.00", tdr),
                    Paragraph("Even Exchange", stat_label),
                ]
        else:
            # ── Sale / Return ──────────────────────────────────────────────
            import json
            p_list = []
            try:
                p_list = json.loads(b.get("products_json") or "[]")
            except Exception:
                pass

            is_return = tx_type == "Return"
            is_jenny  = b.get("jenny") == 1
            price_val = b["price"] or 0.0

            # Build price / balance cell based on transaction type
            td_price_red = ParagraphStyle("tdr_ret", fontName="Helvetica-Bold", fontSize=9,
                                          textColor=RED_C, alignment=TA_RIGHT)
            td_price_mut = ParagraphStyle("tdr_mut", fontName="Helvetica", fontSize=8,
                                          textColor=MUTED, alignment=TA_RIGHT)
            stat_sub     = ParagraphStyle("statsubl", fontName="Helvetica-Oblique", fontSize=7,
                                          textColor=MUTED, alignment=TA_RIGHT)

            if is_return:
                price_cell_val = [
                    Paragraph(f"<b>−{aed(price_val)}</b>", td_price_red),
                    Paragraph("↩ Refunded to Customer", stat_sub),
                ]
            elif is_jenny:
                price_cell_val = [
                    Paragraph(aed(price_val), td_price_mut),
                    Paragraph("(Jenny — not in total)", stat_sub),
                ]
            else:
                price_cell_val = Paragraph(aed(price_val), tdr)

            if p_list:
                details_parts = []
                dta_parts = []
                for p_idx, p in enumerate(p_list):
                    p_brand = (p.get("brand") or "").strip()
                    p_model = (p.get("model") or "").strip()
                    p_dta = (p.get("dta") or "").strip().upper()
                    p_price = p.get("price") or 0.0
                    p_qty = int(p.get("quantity") or 1)
                    qty_str = f" &nbsp;|&nbsp; Qty: {p_qty}"
                    if p_qty > 1:
                        qty_str += f" &nbsp;|&nbsp; Total: {aed(p_price * p_qty)}"
                    details_parts.append(
                        f"<b>Item {p_idx+1}:</b> <b>{p_brand}</b> {get_model_name(p_model)}<br/>"
                        f"<font size='7' color='#6c757d'>&nbsp;&nbsp;DTA: {p_dta} &nbsp;|&nbsp; Price: {aed(p_price)}{qty_str}</font>"
                    )
                    dta_parts.append(p_dta)
                
                details_text = "<br/>".join(details_parts)
                product_cell = Paragraph(details_text, td)
                dta_cell = Paragraph("<br/>".join(dta_parts), tdc)
                bal_cell = price_cell_val
            else:
                brand = (b["brand"] or "").strip()
                model = (b["model"] or "").strip()
                details_text = (
                    f"<b>{brand}</b> {get_model_name(model)}"
                )
                product_cell = Paragraph(details_text, td)
                dta_cell = Paragraph(b["dta"] or "—", tdc)
                bal_cell = price_cell_val

        cust_name = b["customer_name"] or "—"
        if b.get("jenny") == 1:
            cust_name = f"<font color='#be185d'><b>{cust_name} (Jenny)</b></font>"

        row = [
            Paragraph(str(i + 1), tdc),
            Paragraph(cust_name, td),
            product_cell,
            dta_cell,
            bal_cell,
            Paragraph(mop_text(b), td),
            Paragraph(tx_type, tdc),
            Paragraph("Yes" if b["delivery"] else "No", tdc),
            Paragraph(b["note"] or "—", td),
        ]
        table_data.append(row)

    tx_table = Table(table_data, colWidths=COL_WIDTHS, repeatRows=1)

    # Build per-row styling
    ts = [
        # Header row
        ("BACKGROUND",    (0, 0), (-1, 0), BRAND_COLOR),
        ("TEXTCOLOR",     (0, 0), (-1, 0), WHITE),
        ("ALIGN",         (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0), 8),
        ("TOPPADDING",    (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        # All data rows
        ("FONTSIZE",      (0, 1), (-1, -1), 8),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
        ("BOX",           (0, 0), (-1, -1), 0.5, MID_GRAY),
        ("INNERGRID",     (0, 0), (-1, -1), 0.3, MID_GRAY),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [WHITE, ALT_ROW]),
    ]

    # Color-code rows by type
    for i, b in enumerate(bills):
        row_idx = i + 1
        tx = b["transaction_type"]
        if tx == "Exchange":
            ts.append(("BACKGROUND", (0, row_idx), (-1, row_idx), EXCH_BG))
        elif tx == "Return":
            ts.append(("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor("#fff0f0")))
        # Color the Type cell
        type_col = 6
        if tx == "Sale":
            ts.append(("TEXTCOLOR", (type_col, row_idx), (type_col, row_idx), GREEN))
            ts.append(("FONTNAME",  (type_col, row_idx), (type_col, row_idx), "Helvetica-Bold"))
        elif tx == "Return":
            ts.append(("TEXTCOLOR", (type_col, row_idx), (type_col, row_idx), RED_C))
            ts.append(("FONTNAME",  (type_col, row_idx), (type_col, row_idx), "Helvetica-Bold"))
        elif tx == "Exchange":
            ts.append(("TEXTCOLOR", (type_col, row_idx), (type_col, row_idx), BLUE))
            ts.append(("FONTNAME",  (type_col, row_idx), (type_col, row_idx), "Helvetica-Bold"))

    tx_table.setStyle(TableStyle(ts))
    story.append(tx_table)

    # ── Footer note ────────────────────────────────────────────────────────────
    story.append(Spacer(1, 4 * mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=MID_GRAY))
    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph(
        f"Exported by Buyology — {date}  |  Sales: {len(sales) + len(exchanges)}  Returns: {len(returns)}  Exchanges: {len(exchanges)}  |  Total Revenue: {aed(total_rev)}",
        note_style
    ))

    doc.build(story)
    buf.seek(0)

    filename = f"Buyology_Sales_{date}.pdf"
    return send_file(
        buf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )



# ── Marketing PDF Export ──────────────────────────────────────────────────────

@app.route("/api/export-pdf-marketing", methods=["GET"])
@permission_required("marketing")
def api_export_pdf_marketing():
    month = request.args.get("month", "")
    mode  = request.args.get("mode", "platforms")   # "platforms" or "purchase"

    if not month:
        month = _date.today().strftime("%m-%Y")

    # Fetch all bills for the month
    conn = db.get_connection()
    rows = conn.execute(
        "SELECT * FROM bills WHERE date LIKE ? ORDER BY date ASC, id ASC",
        (f"%-{month}",)
    ).fetchall()
    conn.close()
    bills = [dict(r) for r in rows if r["jenny"] != 1]

    # Month label
    mm, yyyy = month.split("-")
    month_names = ["January","February","March","April","May","June",
                   "July","August","September","October","November","December"]
    month_label = f"{month_names[int(mm)-1]} {yyyy}"

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable, Image, KeepTogether
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    except ImportError:
        return jsonify({"error": "reportlab not installed. Run: py -m pip install reportlab"}), 500

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm,
        topMargin=15*mm, bottomMargin=15*mm,
    )

    # ── Colours ───────────────────────────────────────────────────────────────
    BRAND   = colors.HexColor("#0F0A1E")
    PURPLE  = colors.HexColor("#5B3FD4")
    GOLD    = colors.HexColor("#F5A623")
    GREEN   = colors.HexColor("#10b981")
    GREEN_L = colors.HexColor("#d1fae5")
    RED_C   = colors.HexColor("#ef4444")
    RED_L   = colors.HexColor("#fee2e2")
    BLUE    = colors.HexColor("#3b82f6")
    BLUE_L  = colors.HexColor("#dbeafe")
    TEAL    = colors.HexColor("#14b8a6")
    TEAL_L  = colors.HexColor("#ccfbf1")
    AMBER   = colors.HexColor("#f59e0b")
    AMBER_L = colors.HexColor("#fef3c7")
    LIGHT   = colors.HexColor("#f8f7fe")
    BORDER  = colors.HexColor("#e5e2f8")
    MUTED   = colors.HexColor("#6b7280")
    WHITE   = colors.white
    DARK    = colors.HexColor("#1A1330")
    ALT_ROW = colors.HexColor("#F5F3FF")

    # ── Styles ────────────────────────────────────────────────────────────────
    def sty(name, **kw):
        return ParagraphStyle(name, **kw)

    H1  = sty("H1",  fontName="Helvetica-Bold",    fontSize=18, textColor=WHITE,  alignment=TA_LEFT,   leading=22)
    H2  = sty("H2",  fontName="Helvetica-Bold",    fontSize=9,  textColor=GOLD,   alignment=TA_LEFT)
    H2R = sty("H2R", fontName="Helvetica-Bold",    fontSize=9,  textColor=GOLD,   alignment=TA_RIGHT)
    SEC = sty("SEC", fontName="Helvetica-Bold",    fontSize=11, textColor=PURPLE, spaceBefore=8, spaceAfter=3)
    TH  = sty("TH",  fontName="Helvetica-Bold",    fontSize=8,  textColor=WHITE,  alignment=TA_CENTER)
    THL = sty("THL", fontName="Helvetica-Bold",    fontSize=8,  textColor=WHITE,  alignment=TA_LEFT)
    TD  = sty("TD",  fontName="Helvetica",         fontSize=8,  textColor=DARK,   alignment=TA_LEFT)
    TDC = sty("TDC", fontName="Helvetica",         fontSize=8,  textColor=DARK,   alignment=TA_CENTER)
    TDR = sty("TDR", fontName="Helvetica-Bold",    fontSize=8,  textColor=DARK,   alignment=TA_RIGHT)
    GRN = sty("GRN", fontName="Helvetica-Bold",    fontSize=9,  textColor=GREEN,  alignment=TA_RIGHT)
    MUT = sty("MUT", fontName="Helvetica-Oblique", fontSize=7,  textColor=MUTED)
    STV = sty("STV", fontName="Helvetica-Bold",    fontSize=13, textColor=DARK,   alignment=TA_CENTER)
    STL = sty("STL", fontName="Helvetica",         fontSize=7,  textColor=MUTED,  alignment=TA_CENTER)
    INS = sty("INS", fontName="Helvetica-Oblique", fontSize=8.5, textColor=DARK,  leading=13)

    def aed(v):
        try: return f"AED {float(v):,.0f}"
        except: return "AED 0"

    def aed2(v):
        try: return f"AED {float(v):,.2f}"
        except: return "AED 0.00"

    def stat_cell(val, lbl, color=DARK):
        style = ParagraphStyle("sv_dyn", fontName="Helvetica-Bold", fontSize=13,
                               textColor=color, alignment=TA_CENTER)
        return [Paragraph(str(val), style), Paragraph(lbl, STL)]

    # ── Compute global metrics ────────────────────────────────────────────────
    total_tx    = len(bills)
    total_rev   = 0.0
    total_sales = 0
    total_ret   = 0
    total_exch  = 0
    mop_sum     = {"Cash": 0.0, "Card": 0.0, "Tabby": 0.0, "Tamara": 0.0,
                   "Bank Transfer": 0.0, "Pending": 0.0, "Mixed": 0.0}

    for b in bills:
        if b.get("jenny") == 1:
            continue
        tx  = b["transaction_type"]
        amt = 0.0
        sign = -1 if tx == "Return" else 1
        
        if tx == "Sale" or tx == "Return":
            if tx == "Sale":
                total_sales += 1
            else:
                total_ret += 1
            amt = b["price"] or 0.0
            total_rev += sign * amt
        elif tx == "Exchange":
            total_sales += 1
            total_exch  += 1
            bal = b["exch_balance"] or 0.0
            if bal > 0:
                amt = bal
                total_rev += amt

        # MOP sum: subtract for Return, add for Sale/Exchange
        mop = b["payment_mode"]
        if mop == "Mixed":
            mop_sum["Cash"]   += sign * (b["mixed_cash"]   or 0.0)
            mop_sum["Card"]   += sign * (b["mixed_card"]   or 0.0)
            mop_sum["Tabby"]  += sign * (b["mixed_tabby"]  or 0.0)
            mop_sum["Tamara"] += sign * (b["mixed_tamara"] or 0.0)
            mop_sum["Bank Transfer"] += sign * (b["mixed_bank"] or 0.0)
        elif mop in mop_sum:
            mop_sum[mop] += sign * amt

    avg_ticket = (total_rev / total_sales) if total_sales > 0 else 0.0
    total_return_val_mkt = sum(
        (b["price"] or 0.0) for b in bills
        if b["transaction_type"] == "Return" and not b.get("jenny")
    )
    jenny_count_mkt = sum(1 for b in bills if b.get("jenny") == 1)

    # ── Segment data by mode ─────────────────────────────────────────────────
    if mode == "purchase":
        segments = ["In-Store", "Delivery"]
        seg_filter = {
            "In-Store": lambda b: b["delivery"] == 0 or b["delivery"] is False,
            "Delivery": lambda b: b["delivery"] == 1 or b["delivery"] is True,
        }
        seg_colors = [PURPLE, TEAL]
        report_title = "Fulfillment Channel Report"
        report_sub   = "In-Store vs Delivery — Performance Comparison"
    else:
        segments = ["Instagram", "TikTok", "Reference", "Regular Customer"]
        seg_filter = {s: (lambda b, _s=s: b["platform"] == _s) for s in segments}
        seg_colors = [PURPLE, BLUE, AMBER, GREEN]
        report_title = "Marketing Channel Report"
        report_sub   = "Instagram · TikTok · Reference · Regular Customer — Performance Comparison"

    seg_data = {}
    for seg in segments:
        seg_bills = [b for b in bills if seg_filter[seg](b) and b.get("jenny") != 1]
        rev = 0.0; sales = 0; ret = 0; exch = 0; deliv = 0
        daily = {}
        for b in seg_bills:
            tx = b["transaction_type"]
            amt = 0.0
            sign = -1 if tx == "Return" else 1
            if tx == "Sale":
                sales += 1
                amt = b["price"] or 0.0
                rev += amt
            elif tx == "Exchange":
                sales += 1; exch += 1
                bal = b["exch_balance"] or 0.0
                if bal > 0:
                    amt = bal
                    rev += amt
            elif tx == "Return":
                ret += 1
                amt = b["price"] or 0.0
                rev -= amt
            if b["delivery"]: deliv += 1
            day = (b["date"] or "").split("-")[0]
            daily[day] = daily.get(day, 0.0) + (sign * amt)

        seg_data[seg] = {
            "bills": seg_bills, "total": len(seg_bills),
            "sales": sales, "returns": ret, "exchanges": exch,
            "revenue": rev, "delivery": deliv,
            "avg": rev / sales if sales > 0 else 0.0,
            "share": (len(seg_bills) / total_tx * 100) if total_tx > 0 else 0.0,
            "rev_share": (rev / total_rev * 100) if total_rev > 0 else 0.0,
            "daily": daily,
        }

    best_seg = max(segments, key=lambda s: seg_data[s]["revenue"]) if any(seg_data[s]["revenue"] > 0 for s in segments) else segments[0]

    story = []

    # ══════════════════════════════════════════════════════════════════════════
    # 1. HEADER BANNER
    # ══════════════════════════════════════════════════════════════════════════
    logo_path = os.path.join(os.path.dirname(__file__), "static", "logo.png")
    logo_w = None
    if os.path.exists(logo_path):
        try: logo_w = Image(logo_path, width=38*mm, height=8*mm)
        except: pass

    header_left = logo_w or Paragraph("<b>BUYOLOGY</b>", H1)
    hdr = Table([[header_left,
                  Paragraph(f"{report_title} — {month_label}", H2),
                  Paragraph(f"Generated: {_date.today().strftime('%d %B %Y')}", H2R)]],
                colWidths=[55*mm, 90*mm, 30*mm])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), BRAND),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(hdr)
    story.append(Paragraph(f"  {report_sub}", MUT))
    story.append(Spacer(1, 4*mm))

    # ══════════════════════════════════════════════════════════════════════════
    # 2. GLOBAL SUMMARY STATS ROW
    # ══════════════════════════════════════════════════════════════════════════
    story.append(Paragraph("Monthly Overview", SEC))
    story.append(HRFlowable(width="100%", thickness=1.5, color=PURPLE, spaceAfter=3))

    sum_cells = [
        stat_cell(f"{total_tx}", f"Total Transactions\n(incl. {total_ret} returns)"),
        stat_cell(aed(total_rev), "Net Revenue\n(after returns)", GREEN),
        stat_cell(f"{total_sales}", "Sales & Exchanges"),
        stat_cell(
            f"{total_ret}   −{aed(total_return_val_mkt)}",
            "Returns\n(deducted from total)",
            RED_C
        ),
        stat_cell(f"{total_exch}", "Exchanges", BLUE),
        stat_cell(
            f"{aed(avg_ticket)}\n({jenny_count_mkt} Jenny txns excluded)",
            "Avg. Ticket Size"
        ),
    ]
    sum_tbl = Table([sum_cells], colWidths=["16.67%"]*6)
    sum_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), LIGHT),
        ("BOX",           (0, 0), (-1, -1), 0.5, BORDER),
        ("INNERGRID",     (0, 0), (-1, -1), 0.5, BORDER),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("BACKGROUND",    (1, 0), (1, 0),   GREEN_L),
        ("BACKGROUND",    (3, 0), (3, 0),   colors.HexColor("#fff0f0")),
    ]))
    story.append(sum_tbl)
    story.append(Spacer(1, 5*mm))

    # ══════════════════════════════════════════════════════════════════════════
    # 3. CHANNEL / SEGMENT LEADERBOARD TABLE
    # ══════════════════════════════════════════════════════════════════════════
    story.append(Paragraph("Channel Performance Leaderboard", SEC))
    story.append(HRFlowable(width="100%", thickness=1.5, color=PURPLE, spaceAfter=3))

    lb_headers = ["Channel / Platform", "Transactions", "Sales", "Returns", "Revenue",
                  "Avg. Ticket", "Delivery Orders", "Vol. Share", "Rev. Share"]
    lb_widths  = [38*mm, 21*mm, 16*mm, 16*mm, 28*mm, 24*mm, 24*mm, 20*mm, 20*mm]

    lb_data = [[Paragraph(h, TH if i > 0 else THL) for i, h in enumerate(lb_headers)]]

    for idx, seg in enumerate(segments):
        d = seg_data[seg]
        color = seg_colors[idx]
        is_best = seg == best_seg
        row_bg = GREEN_L if is_best else (ALT_ROW if idx % 2 == 0 else WHITE)
        lb_data.append([
            Paragraph(f"<b>{seg}</b>" + (" ★" if is_best else ""), TD),
            Paragraph(str(d["total"]), TDC),
            Paragraph(str(d["sales"]), TDC),
            Paragraph(str(d["returns"]), TDC),
            Paragraph(f"<b>{aed(d['revenue'])}</b>", TDR),
            Paragraph(aed(d["avg"]), TDR),
            Paragraph(f"{d['delivery']} orders", TDC),
            Paragraph(f"{d['share']:.1f}%", TDC),
            Paragraph(f"<b>{d['rev_share']:.1f}%</b>", TDC),
        ])

    lb_tbl = Table(lb_data, colWidths=lb_widths, repeatRows=1)
    lb_ts = [
        ("BACKGROUND",    (0, 0), (-1, 0),   PURPLE),
        ("TEXTCOLOR",     (0, 0), (-1, 0),   WHITE),
        ("FONTNAME",      (0, 0), (-1, 0),   "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1),  8),
        ("VALIGN",        (0, 0), (-1, -1),  "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1),  5),
        ("BOTTOMPADDING", (0, 0), (-1, -1),  5),
        ("LEFTPADDING",   (0, 0), (-1, -1),  5),
        ("RIGHTPADDING",  (0, 0), (-1, -1),  5),
        ("BOX",           (0, 0), (-1, -1),  0.5, BORDER),
        ("INNERGRID",     (0, 0), (-1, -1),  0.3, BORDER),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1),  [WHITE, ALT_ROW]),
    ]
    # Highlight best row
    for i, seg in enumerate(segments):
        if seg == best_seg:
            lb_ts.append(("BACKGROUND", (0, i+1), (-1, i+1), GREEN_L))
    lb_tbl.setStyle(TableStyle(lb_ts))
    story.append(lb_tbl)
    story.append(Spacer(1, 5*mm))

    # ══════════════════════════════════════════════════════════════════════════
    # 4. PAYMENT METHOD BREAKDOWN
    # ══════════════════════════════════════════════════════════════════════════
    story.append(Paragraph("Payment Method (MOP) Breakdown", SEC))
    story.append(HRFlowable(width="100%", thickness=1.5, color=PURPLE, spaceAfter=3))

    mop_order  = ["Cash", "Card", "Tabby", "Tamara", "Bank Transfer", "Pending"]
    mop_bgs    = [GREEN_L, BLUE_L, TEAL_L, colors.HexColor("#fdf2f8"), colors.HexColor("#f3f4f6"), AMBER_L]
    mop_colors = [GREEN, BLUE, TEAL, colors.HexColor("#be185d"), colors.HexColor("#4b5563"), AMBER]

    mop_cells = []
    for i, mop in enumerate(mop_order):
        val = mop_sum.get(mop, 0.0)
        pct = (val / total_rev * 100) if total_rev > 0 else 0.0
        c_style = ParagraphStyle(f"mv{i}", fontName="Helvetica-Bold", fontSize=11,
                                 textColor=mop_colors[i], alignment=TA_CENTER)
        mop_cells.append([
            Paragraph(f"<b>{aed(val)}</b>", c_style),
            Paragraph(f"{mop}\n({pct:.1f}%)", STL),
        ])

    mop_tbl = Table([mop_cells], colWidths=["16.67%"]*6)
    mop_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (0, 0),   GREEN_L),
        ("BACKGROUND",    (1, 0), (1, 0),   BLUE_L),
        ("BACKGROUND",    (2, 0), (2, 0),   TEAL_L),
        ("BACKGROUND",    (3, 0), (3, 0),   colors.HexColor("#fdf2f8")),
        ("BACKGROUND",    (4, 0), (4, 0),   colors.HexColor("#f3f4f6")),
        ("BACKGROUND",    (5, 0), (5, 0),   AMBER_L),
        ("BOX",           (0, 0), (-1, -1), 0.5, BORDER),
        ("INNERGRID",     (0, 0), (-1, -1), 0.5, BORDER),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(mop_tbl)
    story.append(Spacer(1, 5*mm))

    # ══════════════════════════════════════════════════════════════════════════
    # 5. DAILY REVENUE TREND TABLE (per segment)
    # ══════════════════════════════════════════════════════════════════════════
    story.append(Paragraph("Daily Revenue Trend by Channel", SEC))
    story.append(HRFlowable(width="100%", thickness=1.5, color=PURPLE, spaceAfter=3))

    # Collect all active days in month
    all_days = sorted(set(
        b["date"].split("-")[0] for b in bills if b.get("date")
    ), key=lambda d: int(d))

    if all_days:
        trend_headers = ["Day"] + segments + ["Daily Total"]
        trend_widths  = [12*mm] + [int(130 / len(segments))*mm]*len(segments) + [22*mm]
        trend_data    = [[Paragraph(h, TH if i > 0 else THL)
                          for i, h in enumerate(trend_headers)]]

        col_totals = {s: 0.0 for s in segments}
        for day in all_days:
            row = [Paragraph(f"{day}", TDC)]
            day_total = 0.0
            for seg in segments:
                val = seg_data[seg]["daily"].get(day, 0.0)
                col_totals[seg] += val
                day_total += val
                row.append(Paragraph(aed(val) if val > 0 else "—", TDR if val > 0 else TDC))
            row.append(Paragraph(f"<b>{aed(day_total)}</b>", GRN if day_total > 0 else TDC))
            trend_data.append(row)

        # Totals row
        totals_row = [Paragraph("<b>Total</b>", TD)]
        for seg in segments:
            totals_row.append(Paragraph(f"<b>{aed(col_totals[seg])}</b>", TDR))
        totals_row.append(Paragraph(f"<b>{aed(total_rev)}</b>", GRN))
        trend_data.append(totals_row)

        trend_tbl = Table(trend_data, colWidths=trend_widths, repeatRows=1)
        trend_ts = [
            ("BACKGROUND",    (0, 0), (-1, 0),   PURPLE),
            ("TEXTCOLOR",     (0, 0), (-1, 0),   WHITE),
            ("FONTSIZE",      (0, 0), (-1, -1),  7.5),
            ("VALIGN",        (0, 0), (-1, -1),  "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, -1),  4),
            ("BOTTOMPADDING", (0, 0), (-1, -1),  4),
            ("LEFTPADDING",   (0, 0), (-1, -1),  4),
            ("RIGHTPADDING",  (0, 0), (-1, -1),  4),
            ("BOX",           (0, 0), (-1, -1),  0.5, BORDER),
            ("INNERGRID",     (0, 0), (-1, -1),  0.3, BORDER),
            ("ROWBACKGROUNDS",(0, 1), (-1, -2),  [WHITE, ALT_ROW]),
            ("BACKGROUND",    (0, -1), (-1, -1), LIGHT),
            ("FONTNAME",      (0, -1), (-1, -1), "Helvetica-Bold"),
            ("LINEABOVE",     (0, -1), (-1, -1), 1.0, PURPLE),
        ]
        trend_tbl.setStyle(TableStyle(trend_ts))
        story.append(trend_tbl)
    else:
        story.append(Paragraph("No daily revenue data available for this month.", MUT))

    story.append(Spacer(1, 5*mm))

    # ══════════════════════════════════════════════════════════════════════════
    # 6. TOP CUSTOMERS PER CHANNEL
    # ══════════════════════════════════════════════════════════════════════════
    story.append(Paragraph("Top Customers by Channel", SEC))
    story.append(HRFlowable(width="100%", thickness=1.5, color=PURPLE, spaceAfter=3))

    top_cust_tables = []
    col_w = (175 / len(segments)) * mm

    for seg in segments:
        # Aggregate by customer
        cust_map = {}
        for b in seg_data[seg]["bills"]:
            name = (b.get("customer_name") or "Guest").strip() or "Guest"
            tx = b["transaction_type"]
            amt = 0.0
            if tx == "Sale":
                amt = b["price"] or 0.0
            elif tx == "Exchange":
                bal = b["exch_balance"] or 0.0
                if bal > 0: amt = bal
            if name not in cust_map:
                cust_map[name] = {"count": 0, "revenue": 0.0}
            cust_map[name]["count"] += 1
            cust_map[name]["revenue"] += amt

        top5 = sorted(cust_map.items(), key=lambda x: x[1]["revenue"], reverse=True)[:5]

        cust_hdr = [Paragraph(f"<b>{seg}</b>", THL), Paragraph("Txns", TH), Paragraph("Revenue", TH)]
        cust_rows = [cust_hdr]
        for cname, cd in top5:
            display = cname[:22] + "…" if len(cname) > 22 else cname
            cust_rows.append([
                Paragraph(display, TD),
                Paragraph(str(cd["count"]), TDC),
                Paragraph(aed(cd["revenue"]), TDR),
            ])
        if not top5:
            cust_rows.append([Paragraph("No data", MUT), Paragraph("—", TDC), Paragraph("—", TDC)])

        ct = Table(cust_rows, colWidths=[col_w*0.52, col_w*0.18, col_w*0.30])
        ct.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),   PURPLE),
            ("TEXTCOLOR",     (0, 0), (-1, 0),   WHITE),
            ("FONTSIZE",      (0, 0), (-1, -1),  7.5),
            ("VALIGN",        (0, 0), (-1, -1),  "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, -1),  4),
            ("BOTTOMPADDING", (0, 0), (-1, -1),  4),
            ("LEFTPADDING",   (0, 0), (-1, -1),  4),
            ("RIGHTPADDING",  (0, 0), (-1, -1),  4),
            ("BOX",           (0, 0), (-1, -1),  0.5, BORDER),
            ("INNERGRID",     (0, 0), (-1, -1),  0.3, BORDER),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1),  [WHITE, ALT_ROW]),
        ]))
        top_cust_tables.append(ct)

    if top_cust_tables:
        # side by side
        combined = Table([top_cust_tables], colWidths=[col_w]*len(segments))
        combined.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "TOP"), ("LEFTPADDING", (0,0), (-1,-1), 2), ("RIGHTPADDING", (0,0), (-1,-1), 2)]))
        story.append(combined)

    story.append(Spacer(1, 5*mm))

    # ══════════════════════════════════════════════════════════════════════════
    # 7. EXECUTIVE SUMMARY BRIEF
    # ══════════════════════════════════════════════════════════════════════════
    story.append(Paragraph("Executive Summary", SEC))
    story.append(HRFlowable(width="100%", thickness=1.5, color=PURPLE, spaceAfter=3))

    best_d = seg_data[best_seg]
    high_ticket_seg = max(segments, key=lambda s: seg_data[s]["avg"]) if any(seg_data[s]["avg"] > 0 for s in segments) else best_seg
    total_deliveries = sum(seg_data[s]["delivery"] for s in segments)
    del_pct = (total_deliveries / total_tx * 100) if total_tx > 0 else 0.0

    brief = (
        f"During <b>{month_label}</b>, Buyology recorded a total of <b>{total_tx}</b> transactions "
        f"generating gross revenue of <b>{aed2(total_rev)}</b> at an average ticket size of <b>{aed2(avg_ticket)}</b>. "
        f"<br/><br/>"
        f"The top-performing channel was <b>{best_seg}</b>, which contributed <b>{best_d['total']}</b> transactions "
        f"({best_d['share']:.1f}% of total volume) and produced <b>{aed2(best_d['revenue'])}</b> in revenue "
        f"({best_d['rev_share']:.1f}% of total revenue). "
        f"<br/><br/>"
        f"<b>{high_ticket_seg}</b> recorded the highest average ticket size at <b>{aed2(seg_data[high_ticket_seg]['avg'])}</b>, "
        f"indicating premium product purchases from that channel. "
        f"Delivery accounted for <b>{total_deliveries}</b> orders ({del_pct:.1f}% of all transactions). "
        f"<br/><br/>"
        f"Returns: <b>{total_ret}</b> | Exchanges: <b>{total_exch}</b>. "
        f"Cash collections: <b>{aed2(mop_sum['Cash'])}</b> | Card: <b>{aed2(mop_sum['Card'])}</b> | "
        f"Tabby: <b>{aed2(mop_sum['Tabby'])}</b> | Tamara: <b>{aed2(mop_sum['Tamara'])}</b>."
    )

    brief_tbl = Table([[Paragraph(brief, INS)]], colWidths=["100%"])
    brief_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), LIGHT),
        ("BOX",           (0, 0), (-1, -1), 1.0, PURPLE),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING",   (0, 0), (-1, -1), 12),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 12),
    ]))
    story.append(brief_tbl)
    story.append(Spacer(1, 4*mm))

    # ── Footer ────────────────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=BORDER))
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        f"Buyology Marketing Report — {month_label}  |  Mode: {mode.capitalize()}  |  "
        f"Total Revenue: {aed2(total_rev)}  |  Transactions: {total_tx}  |  "
        f"Generated: {_date.today().strftime('%d %B %Y')}",
        MUT
    ))

    doc.build(story)
    buf.seek(0)

    filename = f"Buyology_Marketing_{month_label.replace(' ', '_')}_{mode}.pdf"
    return send_file(
        buf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename,
    )


# ── Inventory Endpoints ──────────────────────────────────────────────────────

@app.route("/api/inventory", methods=["GET"])
@permission_required("inventory")
def api_inventory_list():
    status = request.args.get("status")
    date = request.args.get("date")
    month = request.args.get("month")
    search = request.args.get("search")
    try:
        units = db.get_inventory_units(status=status, date=date, month=month, search=search)
        return jsonify(units)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/inventory/add", methods=["POST"])
@permission_required("inventory")
def api_inventory_add():
    data = request.get_json()
    if not data or not data.get("dta"):
        return jsonify({"error": "DTA code is required"}), 400
        
    dta = data["dta"].strip().upper()
    status = data.get("status", "PURCHASED").strip()
    brand = data.get("brand", "").strip()
    model = data.get("model", "").strip()
    price = float(data.get("price") or 0.0)
    date = data.get("date")
    notes = data.get("notes", "").strip()
    quantity = int(data.get("quantity") or 1)
    
    try:
        res = db.add_inventory_unit(dta, status, brand, model, price, date, notes, quantity)
        return jsonify(res)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/inventory/move", methods=["POST"])
@permission_required("inventory")
def api_inventory_move():
    data = request.get_json()
    if not data or not data.get("dta") or not data.get("to_status"):
        return jsonify({"error": "DTA code and to_status are required"}), 400
        
    dta = data["dta"].strip().upper()
    from_status = data.get("from_status", "STOCK").strip()
    to_status = data["to_status"].strip()
    date = data.get("date")
    notes = data.get("notes", "").strip()
    quantity = int(data.get("quantity") or 1)
    
    try:
        res = db.move_inventory_unit(dta, from_status, to_status, quantity, date, notes)
        return jsonify(res)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/inventory/adjust_quantity", methods=["POST"])
@permission_required("inventory")
def api_inventory_adjust_quantity():
    data = request.get_json()
    if not data or not data.get("dta") or not data.get("status") or "delta" not in data:
        return jsonify({"error": "DTA, status, and delta are required"}), 400
        
    dta = data["dta"].strip().upper()
    status = data["status"].strip()
    delta = int(data["delta"])
    
    try:
        res = db.adjust_inventory_qty(dta, status, delta)
        return jsonify(res)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/inventory/upgrade", methods=["POST"])
@permission_required("inventory")
def api_inventory_upgrade():
    data = request.get_json()
    if not data or not data.get("dta") or not data.get("brand") or not data.get("model"):
        return jsonify({"error": "DTA, brand, and model are required"}), 400
        
    dta = data["dta"].strip().upper()
    brand = data["brand"].strip()
    model = data["model"].strip()
    price = float(data.get("price") or 0.0)
    date = data.get("date")
    notes = data.get("notes", "").strip()
    
    try:
        res = db.upgrade_inventory_unit(dta, brand, model, price, date, notes)
        return jsonify(res)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/inventory/check/<dta>", methods=["GET"])
@permission_required(["inventory", "deliveries", "bills"])
def api_inventory_check(dta):
    try:
        dta = dta.strip().upper()
        unit = db.get_inventory_unit(dta)
        product = db.get_product(dta)
        
        if unit:
            return jsonify({
                "exists": True,
                "in_inventory": True,
                "status": unit["status"],
                "brand": unit["brand"],
                "model": unit["model"],
                "price": unit["price"],
                "notes": unit["notes"]
            })
        elif product:
            return jsonify({
                "exists": True,
                "in_inventory": False,
                "status": None,
                "brand": product["brand"],
                "model": product["model"],
                "price": product["price"],
                "notes": ""
            })
        else:
            return jsonify({
                "exists": False
            })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/inventory/logs", methods=["GET"])
@permission_required("inventory")
def api_inventory_logs():
    dta = request.args.get("dta")
    date = request.args.get("date")
    month = request.args.get("month")
    try:
        logs = db.get_inventory_logs(dta=dta, date=date, month=month)
        return jsonify(logs)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Inventory Report Export APIs ──────────────────────────────────────────────

@app.route("/api/inventory/export/excel", methods=["GET"])
@permission_required("inventory")
def api_inventory_export_excel():
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500

    try:
        status_filter = request.args.get("status")
        date_filter = request.args.get("date")
        month_filter = request.args.get("month")
        
        status_labels = {
            "PURCHASED": "Purchased Intake Queue",
            "QC": "QC Section (Testing)",
            "CLEANED_READY": "Cleaned & Ready",
            "GONE_OUT": "Gone for Outsource",
            "IN_HOUSE_REPAIR": "In-house Repair",
            "SUPPLIER_WARRANTY": "Supplier Warranty",
            "STOCK": "In Stock",
            "NOON": "Noon Branch",
            "REGION_SAUDI": "Saudi Arabia Branch",
            "REGION_QATAR": "Qatar Branch",
            "REGION_OMAN": "Oman Branch",
            "SOLD": "Sold Units",
            "REPAIRS": "Repairs & Warranty",
            "ALL": "All Sections (Consolidated)"
        }
        
        if status_filter == "REPAIRS":
            units = db.get_inventory_units(date=date_filter, month=month_filter)
            units = [u for u in units if u["status"] in ("IN_HOUSE_REPAIR", "SUPPLIER_WARRANTY")]
        elif status_filter and status_filter != "ALL":
            units = db.get_inventory_units(status=status_filter, date=date_filter, month=month_filter)
        else:
            units = db.get_inventory_units(date=date_filter, month=month_filter)
            
        status_order = [
            "PURCHASED", "QC", "CLEANED_READY", "GONE_OUT",
            "IN_HOUSE_REPAIR", "SUPPLIER_WARRANTY", "STOCK", "NOON",
            "REGION_SAUDI", "REGION_QATAR", "REGION_OMAN", "SOLD"
        ]
        
        filter_desc = "All Stock"
        if status_filter and status_filter != "ALL":
            filter_desc = status_labels.get(status_filter, status_filter)
        if date_filter:
            filter_desc += f" on {date_filter}"
        elif month_filter:
            filter_desc += f" in {month_filter}"
            
        safe_filter_desc = filter_desc.replace(" / ", "_").replace(" ", "_").replace(":", "").replace("-", "_")
        
        counts = {s: 0 for s in status_order}
        values = {s: 0.0 for s in status_order}
        for u in units:
            stat = u["status"]
            if stat in counts:
                counts[stat] += 1
                values[stat] += float(u.get("price") or 0.0)
                
        wb = openpyxl.Workbook()
        
        # Sheet 1: Summary Dashboard
        ws1 = wb.active
        ws1.title = "Summary Dashboard"
        
        blue_fill = PatternFill("solid", fgColor="1E3A5F")
        yellow_fill = PatternFill("solid", fgColor="FFD700")
        alt_fill = PatternFill("solid", fgColor="EBF2FF")
        white_bold_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        header_font = Font(bold=True, name="Calibri", size=13)
        bold_font = Font(bold=True, name="Calibri", size=11)
        regular_font = Font(name="Calibri", size=11)
        
        thin_border = Border(
            left=Side(style="thin", color="D3D3D3"),
            right=Side(style="thin", color="D3D3D3"),
            top=Side(style="thin", color="D3D3D3"),
            bottom=Side(style="thin", color="D3D3D3")
        )
        
        ws1.merge_cells("A1:C1")
        title_cell = ws1["A1"]
        title_cell.value = f"Buyology Inventory Summary — {filter_desc} ({today_str()})"
        title_cell.fill = yellow_fill
        title_cell.font = header_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[1].height = 30
        
        headers = ["Inventory Status / Section", "Laptops (Qty)", "Estimated Value (AED)"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws1.cell(row=2, column=col_idx, value=h)
            cell.fill = blue_fill
            cell.font = white_bold_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws1.row_dimensions[2].height = 25
        
        current_row = 3
        active_qty_total = 0
        active_val_total = 0.0
        for s in status_order:
            lbl = status_labels[s]
            qty = counts[s]
            val = values[s]
            
            if s != "SOLD":
                active_qty_total += qty
                active_val_total += val
                
            row_cells = [
                ws1.cell(row=current_row, column=1, value=lbl),
                ws1.cell(row=current_row, column=2, value=qty),
                ws1.cell(row=current_row, column=3, value=val)
            ]
            
            row_cells[0].alignment = Alignment(horizontal="left", vertical="center")
            row_cells[1].alignment = Alignment(horizontal="right", vertical="center")
            row_cells[2].alignment = Alignment(horizontal="right", vertical="center")
            
            row_cells[2].number_format = '#,##0.00'
            
            if current_row % 2 == 1:
                for c in row_cells:
                    c.fill = alt_fill
                    
            for c in row_cells:
                c.font = regular_font
                c.border = thin_border
                
            current_row += 1
            
        total_cells = [
            ws1.cell(row=current_row, column=1, value="Grand Total (Active Staging & Stock)"),
            ws1.cell(row=current_row, column=2, value=active_qty_total),
            ws1.cell(row=current_row, column=3, value=active_val_total)
        ]
        
        for c in total_cells:
            c.font = bold_font
            c.fill = yellow_fill
            c.border = thin_border
            
        total_cells[0].alignment = Alignment(horizontal="left", vertical="center")
        total_cells[1].alignment = Alignment(horizontal="right", vertical="center")
        total_cells[2].alignment = Alignment(horizontal="right", vertical="center")
        total_cells[2].number_format = '#,##0.00'
        
        ws1.column_dimensions["A"].width = 30
        ws1.column_dimensions["B"].width = 15
        ws1.column_dimensions["C"].width = 25
        
        # Sheet 2: Detailed Inventory List
        ws2 = wb.create_sheet(title="Detailed Inventory")
        
        ws2.merge_cells("A1:H1")
        title_cell2 = ws2["A1"]
        title_cell2.value = f"Detailed Inventory List — {filter_desc} ({today_str()})"
        title_cell2.fill = yellow_fill
        title_cell2.font = header_font
        title_cell2.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 30
        
        detail_headers = ["DTA Code", "Brand", "Model", "Specifications", "Price (AED)", "Current Status", "Last Transition Date", "Notes / Details"]
        for col_idx, h in enumerate(detail_headers, 1):
            cell = ws2.cell(row=2, column=col_idx, value=h)
            cell.fill = blue_fill
            cell.font = white_bold_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws2.row_dimensions[2].height = 25
        
        detail_row = 3
        for u in units:
            dta = u["dta"]
            brand = u["brand"]
            model_full = u["model"] or ""
            
            model_parts = model_full.split("|")
            model = model_parts[0].strip()
            specs = model_parts[1].strip() if len(model_parts) > 1 else ""
            
            price = float(u.get("price") or 0.0)
            status_lbl = status_labels.get(u["status"], u["status"])
            date = u["status_date"]
            notes = u["notes"] or ""
            
            cells = [
                ws2.cell(row=detail_row, column=1, value=dta),
                ws2.cell(row=detail_row, column=2, value=brand),
                ws2.cell(row=detail_row, column=3, value=model),
                ws2.cell(row=detail_row, column=4, value=specs),
                ws2.cell(row=detail_row, column=5, value=price),
                ws2.cell(row=detail_row, column=6, value=status_lbl),
                ws2.cell(row=detail_row, column=7, value=date),
                ws2.cell(row=detail_row, column=8, value=notes)
            ]
            
            cells[0].alignment = Alignment(horizontal="center", vertical="center")
            cells[1].alignment = Alignment(horizontal="left", vertical="center")
            cells[2].alignment = Alignment(horizontal="left", vertical="center")
            cells[3].alignment = Alignment(horizontal="left", vertical="center")
            cells[4].alignment = Alignment(horizontal="right", vertical="center")
            cells[5].alignment = Alignment(horizontal="center", vertical="center")
            cells[6].alignment = Alignment(horizontal="center", vertical="center")
            cells[7].alignment = Alignment(horizontal="left", vertical="center")
            
            cells[4].number_format = '#,##0.00'
            
            if detail_row % 2 == 1:
                for c in cells:
                    c.fill = alt_fill
                    
            for c in cells:
                c.font = regular_font
                c.border = thin_border
                
            detail_row += 1
            
        widths = [15, 12, 18, 25, 14, 22, 15, 30]
        for col_idx, w in enumerate(widths, 1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w
            
        ws2.freeze_panes = "A3"
        
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"Buyology_Inventory_{safe_filter_desc}_{today_str()}.xlsx"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/inventory/export/pdf", methods=["GET"])
@permission_required("inventory")
def api_inventory_export_pdf():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    except ImportError:
        return jsonify({"error": "reportlab not installed. Run: py -m pip install reportlab"}), 500

    try:
        status_filter = request.args.get("status")
        date_filter = request.args.get("date")
        month_filter = request.args.get("month")
        
        status_labels = {
            "PURCHASED": "Purchased Intake Queue",
            "QC": "QC Section (Testing)",
            "CLEANED_READY": "Cleaned & Ready",
            "GONE_OUT": "Gone for Outsource",
            "IN_HOUSE_REPAIR": "In-house Repair",
            "SUPPLIER_WARRANTY": "Supplier Warranty",
            "STOCK": "In Stock",
            "NOON": "Noon Branch",
            "REGION_SAUDI": "Saudi Arabia Branch",
            "REGION_QATAR": "Qatar Branch",
            "REGION_OMAN": "Oman Branch",
            "SOLD": "Sold Units",
            "REPAIRS": "Repairs & Warranty",
            "ALL": "All Sections (Consolidated)"
        }
        
        if status_filter == "REPAIRS":
            units = db.get_inventory_units(date=date_filter, month=month_filter)
            units = [u for u in units if u["status"] in ("IN_HOUSE_REPAIR", "SUPPLIER_WARRANTY")]
        elif status_filter and status_filter != "ALL":
            units = db.get_inventory_units(status=status_filter, date=date_filter, month=month_filter)
        else:
            units = db.get_inventory_units(date=date_filter, month=month_filter)

        status_order = [
            "PURCHASED", "QC", "CLEANED_READY", "GONE_OUT",
            "IN_HOUSE_REPAIR", "SUPPLIER_WARRANTY", "STOCK", "NOON",
            "REGION_SAUDI", "REGION_QATAR", "REGION_OMAN", "SOLD"
        ]
        
        filter_desc = "All Stock"
        if status_filter and status_filter != "ALL":
            filter_desc = status_labels.get(status_filter, status_filter)
        if date_filter:
            filter_desc += f" on {date_filter}"
        elif month_filter:
            filter_desc += f" in {month_filter}"
            
        safe_filter_desc = filter_desc.replace(" / ", "_").replace(" ", "_").replace(":", "").replace("-", "_")
        
        counts = {s: 0 for s in status_order}
        values = {s: 0.0 for s in status_order}
        for u in units:
            stat = u["status"]
            if stat in counts:
                counts[stat] += 1
                values[stat] += float(u.get("price") or 0.0)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=A4,
            leftMargin=12 * mm,
            rightMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
        )

        BRAND_COLOR = colors.HexColor("#402F75")
        DARK        = colors.HexColor("#1F1B2C")
        MID_GRAY    = colors.HexColor("#dee2e6")
        WHITE       = colors.white
        MUTED       = colors.HexColor("#6c757d")
        ORANGE      = colors.HexColor("#FCBC12")
        ALT_ROW     = colors.HexColor("#F8F7FA")

        title_style = ParagraphStyle(
            "title", fontName="Helvetica-Bold", fontSize=18,
            textColor=WHITE, alignment=TA_LEFT, leading=22,
        )
        sub_style = ParagraphStyle(
            "sub", fontName="Helvetica-Bold", fontSize=9,
            textColor=ORANGE, alignment=TA_LEFT,
        )
        section_style = ParagraphStyle(
            "section", fontName="Helvetica-Bold", fontSize=12,
            textColor=BRAND_COLOR, spaceBefore=10, spaceAfter=6,
        )
        th_style = ParagraphStyle(
            "th", fontName="Helvetica-Bold", fontSize=8,
            textColor=WHITE, alignment=TA_CENTER,
        )
        td = ParagraphStyle("td", fontName="Helvetica", fontSize=8, textColor=DARK, alignment=TA_LEFT)
        tdc = ParagraphStyle("tdc", fontName="Helvetica", fontSize=8, textColor=DARK, alignment=TA_CENTER)
        tdr = ParagraphStyle("tdr", fontName="Helvetica-Bold", fontSize=8, textColor=DARK, alignment=TA_RIGHT)

        story = []

        header_data = [
            [Paragraph("BUYOLOGY — INVENTORY REPORT", title_style)],
            [Paragraph(f"Generated on {today_str()}  |  {filter_desc}", sub_style)]
        ]
        header_table = Table(header_data, colWidths=[186 * mm])
        header_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), BRAND_COLOR),
            ('PADDING', (0,0), (-1,-1), 8),
            ('BOTTOMPADDING', (0,1), (-1,-1), 10),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 4 * mm))

        story.append(Paragraph("Inventory Summary Dashboard", section_style))
        story.append(HRFlowable(width="100%", thickness=1, color=MID_GRAY, spaceAfter=3 * mm))

        summary_data = [[
            Paragraph("Status / Section", th_style),
            Paragraph("Laptops (Qty)", th_style),
            Paragraph("Estimated Value (AED)", th_style)
        ]]

        active_qty_total = 0
        active_val_total = 0.0
        for s in status_order:
            lbl = status_labels[s]
            qty = counts[s]
            val = values[s]
            if s != "SOLD":
                active_qty_total += qty
                active_val_total += val

            summary_data.append([
                Paragraph(lbl, td),
                Paragraph(str(qty), tdc),
                Paragraph(f"{val:,.2f}", tdr)
            ])

        summary_data.append([
            Paragraph("Grand Total (Active Staging & Stock)", ParagraphStyle("td_bold", parent=td, fontName="Helvetica-Bold")),
            Paragraph(str(active_qty_total), ParagraphStyle("tdc_bold", parent=tdc, fontName="Helvetica-Bold")),
            Paragraph(f"{active_val_total:,.2f}", ParagraphStyle("tdr_bold", parent=tdr, fontName="Helvetica-Bold"))
        ])

        summary_table = Table(summary_data, colWidths=[96 * mm, 35 * mm, 55 * mm])
        summary_table_style = [
            ('BACKGROUND', (0,0), (-1,0), BRAND_COLOR),
            ('GRID', (0,0), (-1,-1), 0.5, MID_GRAY),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 5),
            ('BACKGROUND', (0,-1), (-1,-1), ORANGE),
        ]
        
        for i in range(1, len(summary_data) - 1):
            if i % 2 == 1:
                summary_table_style.append(('BACKGROUND', (0, i), (-1, i), ALT_ROW))
                
        summary_table.setStyle(TableStyle(summary_table_style))
        story.append(summary_table)
        story.append(Spacer(1, 6 * mm))

        story.append(Paragraph("Detailed Laptop Status Report", section_style))
        story.append(HRFlowable(width="100%", thickness=1, color=MID_GRAY, spaceAfter=4 * mm))

        units_by_status = {s: [] for s in status_order}
        for u in units:
            stat = u["status"]
            if stat in units_by_status:
                units_by_status[stat].append(u)

        for s in status_order:
            status_units = units_by_status[s]
            if not status_units:
                continue

            story.append(Paragraph(f"{status_labels[s]} ({len(status_units)} Laptops)", ParagraphStyle("sub_sec", parent=section_style, fontSize=10, textColor=DARK, spaceBefore=4, spaceAfter=2)))
            
            detail_data = [[
                Paragraph("DTA Code", th_style),
                Paragraph("Brand / Model", th_style),
                Paragraph("Price (AED)", th_style),
                Paragraph("Updated Date", th_style),
                Paragraph("Notes / Customization", th_style)
            ]]

            for u in status_units:
                model_name = (u["model"] or "").split("|")[0].strip()
                brand_model = f"<b>{u['brand']}</b> {model_name}"
                price = float(u.get("price") or 0.0)
                notes = u["notes"] or "—"
                date = u["status_date"]

                detail_data.append([
                    Paragraph(u["dta"], tdc),
                    Paragraph(brand_model, td),
                    Paragraph(f"{price:,.2f}", tdr),
                    Paragraph(date, tdc),
                    Paragraph(notes, ParagraphStyle("td_notes", parent=td, fontSize=7.5, leading=9))
                ])

            detail_table = Table(detail_data, colWidths=[24 * mm, 42 * mm, 24 * mm, 24 * mm, 72 * mm])
            detail_table_style = [
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#5D3FD3")),
                ('GRID', (0,0), (-1,-1), 0.5, MID_GRAY),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('PADDING', (0,0), (-1,-1), 4),
            ]
            for i in range(1, len(detail_data)):
                if i % 2 == 1:
                    detail_table_style.append(('BACKGROUND', (0, i), (-1, i), ALT_ROW))
            detail_table.setStyle(TableStyle(detail_table_style))
            story.append(detail_table)
            story.append(Spacer(1, 4 * mm))

        doc.build(story)
        buf.seek(0)
        
        return send_file(
            buf,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"Buyology_Inventory_{safe_filter_desc}_{today_str()}.pdf"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Display Pieces API ────────────────────────────────────────────────────────


@app.route("/api/display", methods=["GET"])
@permission_required("display")
def api_display_list():
    date = request.args.get("date")
    return jsonify(db.get_all_display_pieces(date))


@app.route("/api/display", methods=["POST"])
@permission_required("display")
def api_display_add():
    data = request.get_json()
    if not data or not data.get("dta"):
        return jsonify({"error": "DTA is required"}), 400
    
    dta   = data["dta"].strip().upper()
    brand = data.get("brand", "").strip()
    model = data.get("model", "").strip()
    force = bool(data.get("force", False))
    date  = data.get("date")
    qty   = int(data.get("quantity") or 1)
    
    try:
        res = db.add_display_piece(dta, brand, model, force_increment=force, date=date, qty=qty)
        return jsonify(res)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/display/<dta>", methods=["DELETE"])
@permission_required("display")
def api_display_delete(dta):
    date = request.args.get("date")
    qty  = int(request.args.get("quantity") or 1)
    try:
        db.decrement_display_piece(dta, date, qty=qty)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/display/check/<dta>", methods=["GET"])
@permission_required("display")
def api_display_check(dta):
    date = request.args.get("date")
    is_display = db.check_display_piece(dta, date)
    return jsonify({"display": is_display})


@app.route("/api/display/rollover", methods=["POST"])
@permission_required("display")
def api_display_rollover():
    data = request.get_json()
    if not data or not data.get("source_date") or not data.get("target_date"):
        return jsonify({"error": "source_date and target_date are required"}), 400
        
    source_date = data["source_date"].strip()
    target_date = data["target_date"].strip()
    
    try:
        res = db.rollover_display_pieces(source_date, target_date)
        return jsonify(res)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/display/export-pdf", methods=["GET"])
@permission_required("display")
def api_display_export_pdf():
    date = request.args.get("date", today_str())
    pieces = db.get_all_display_pieces(date)

    from datetime import datetime as _datetime
    try:
        pdf_date_label = _datetime.strptime(date, "%d-%m-%Y").strftime("%d %b %Y")
    except Exception:
        pdf_date_label = date

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, Image
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    except ImportError:
        return jsonify({"error": "reportlab not installed. Run: py -m pip install reportlab"}), 500

    buf = io.BytesIO()

    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    BRAND_COLOR  = colors.HexColor("#402F75")
    DARK         = colors.HexColor("#1F1B2C")
    LIGHT_GRAY   = colors.HexColor("#f5f7fa")
    MID_GRAY     = colors.HexColor("#dee2e6")
    ALT_ROW      = colors.HexColor("#F8F7FA")
    WHITE        = colors.white
    MUTED        = colors.HexColor("#6c757d")

    title_style = ParagraphStyle(
        "title", fontName="Helvetica-Bold", fontSize=16,
        textColor=WHITE, alignment=TA_LEFT, leading=20,
    )
    sub_style = ParagraphStyle(
        "sub", fontName="Helvetica-Bold", fontSize=9,
        textColor=colors.HexColor("#FCBC12"), alignment=TA_LEFT,
    )
    sub_right_style = ParagraphStyle(
        "sub_right", parent=sub_style, alignment=TA_RIGHT,
    )
    th_style = ParagraphStyle(
        "th", fontName="Helvetica-Bold", fontSize=8.5,
        textColor=WHITE, alignment=TA_LEFT,
    )
    th_center = ParagraphStyle(
        "thc", fontName="Helvetica-Bold", fontSize=8.5,
        textColor=WHITE, alignment=TA_CENTER,
    )
    
    td  = ParagraphStyle("td",  fontName="Helvetica", fontSize=8, textColor=DARK, alignment=TA_LEFT)
    td_bold = ParagraphStyle("tdb", fontName="Helvetica-Bold", fontSize=8, textColor=BRAND_COLOR, alignment=TA_LEFT)
    tdc = ParagraphStyle("tdc", fontName="Helvetica", fontSize=8, textColor=DARK, alignment=TA_CENTER)
    td_muted  = ParagraphStyle("tdm", fontName="Helvetica", fontSize=7.5, textColor=MUTED, alignment=TA_LEFT)

    story = []

    # ── Header ────────────────────────────────────────────────────────────────
    logo_path = os.path.join(os.path.dirname(__file__), "static", "logo.png")
    logo_widget = None
    if os.path.exists(logo_path):
        try:
            logo_widget = Image(logo_path, width=32*mm, height=7*mm)
        except Exception:
            pass

    header_left = logo_widget if logo_widget else Paragraph("<b>BUYOLOGY</b>", title_style)

    header_data = [[
        header_left,
        Paragraph("Display Pieces Registry", sub_style),
        Paragraph(f"Date: {pdf_date_label}", sub_right_style),
    ]]
    header_table = Table(header_data, colWidths=[55*mm, 75*mm, 50*mm])
    header_table.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, -1), BRAND_COLOR),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
        ("TOPPADDING",   (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 8),
        ("ALIGN",        (2, 0), (2, 0),   "RIGHT"),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 5 * mm))

    # ── Stats Summary ─────────────────────────────────────────────────────────
    total_qty = sum(p.get("quantity", 1) or 1 for p in pieces)
    stat_bold  = ParagraphStyle("statb", fontName="Helvetica-Bold", fontSize=11, alignment=TA_CENTER)
    stat_lbl2  = ParagraphStyle("statl", fontName="Helvetica", fontSize=8, textColor=MUTED, alignment=TA_CENTER)

    def stat_cell(value, label):
        return [Paragraph(str(value), stat_bold), Paragraph(label, stat_lbl2)]

    summary_data = [[
        stat_cell(len(pieces), "Unique Models on Display"),
        stat_cell(total_qty, "Total Display Units Qty")
    ]]
    summary_table = Table(summary_data, colWidths=[90*mm, 90*mm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, -1), LIGHT_GRAY),
        ("BOX",          (0, 0), (-1, -1), 1, MID_GRAY),
        ("INNERGRID",    (0, 0), (-1, -1), 0.5, MID_GRAY),
        ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",   (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 8),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 5 * mm))

    # ── Table ─────────────────────────────────────────────────────────────────
    col_widths = [35*mm, 133*mm, 12*mm]
    
    table_data = [[
        Paragraph("DTA Code", th_style),
        Paragraph("Laptop Description & Specifications", th_style),
        Paragraph("Qty", th_center),
    ]]

    for idx, p in enumerate(pieces):
        dta = p.get("dta", "")
        brand = p.get("brand", "").strip().upper()
        
        full_model = p.get("model", "")
        model_parts = [m.strip() for m in full_model.split("|")] if full_model else []
        if model_parts:
            model_name = model_parts[0]
            # Strip brand from model name if present to avoid duplication
            if brand and model_name.lower().startswith(brand.lower()):
                model_name = model_name[len(brand):].strip()
            
            model_name_upper = model_name.upper()
            specs_list = model_parts[1:]
            
            desc_parts = [f"{brand} {model_name_upper}"] if brand else [model_name_upper]
            desc_parts.extend(specs_list)
            desc_text = " | ".join(desc_parts)
        else:
            desc_text = brand if brand else "—"
            
        qty = p.get("quantity", 1) or 1
        
        table_data.append([
            Paragraph(f"<b>{dta}</b>", td_bold),
            Paragraph(desc_text, td),
            Paragraph(str(qty), tdc),
        ])

    table_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_COLOR),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.5, MID_GRAY),
    ])

    for i in range(1, len(table_data)):
        bg = ALT_ROW if i % 2 == 1 else WHITE
        table_style.add("BACKGROUND", (0, i), (-1, i), bg)

    t = Table(table_data, colWidths=col_widths)
    t.setStyle(table_style)
    story.append(t)

    doc.build(story)
    buf.seek(0)
    
    filename = f"Display_Pieces_Report_{_date.today().strftime('%Y-%m-%d')}.pdf"
    response = send_file(
        buf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename
    )
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

@app.route("/download-cert", methods=["GET"])
def api_download_cert():
    try:
        return send_file("cert.crt", mimetype="application/x-x509-ca-cert", as_attachment=True, download_name="buyology-secure.crt")
    except Exception as e:
        return jsonify({"error": str(e)}), 404


# ── Deliveries API ────────────────────────────────────────────────────────────

@app.route("/api/deliveries", methods=["GET"])
@permission_required("deliveries")
def api_deliveries_list():
    date = request.args.get("date")
    status = request.args.get("status")
    month = request.args.get("month")  # MM-YYYY format
    return jsonify(db.get_all_deliveries(date=date, status=status, month=month))


@app.route("/api/deliveries/dates", methods=["GET"])
@permission_required("deliveries")
def api_deliveries_dates():
    """Return all distinct delivery dates — used to build month selector."""
    return jsonify(db.get_all_delivery_dates())


@app.route("/api/deliveries", methods=["POST"])
@permission_required("deliveries")
def api_deliveries_add():
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400
        
    required = ["customer_name", "place", "phone", "products_json", "price", "delivery_by", "payment_mode"]
    for f in required:
        val = data.get(f)
        if val is None or (isinstance(val, str) and not val.strip()):
            return jsonify({"error": f"Field '{f}' is required"}), 400

    # Date defaults to today if not provided
    date_val = data.get("date") or today_str()
    
    # Extract DTA list automatically from products_json
    import json
    try:
        p_list = json.loads(data["products_json"])
        dtas = [p.get("dta", "").strip().upper() for p in p_list if p.get("dta")]
        dta_list = ", ".join(dtas)
    except Exception:
        dta_list = ""

    try:
        res = db.add_delivery(
            date=date_val,
            customer_name=data["customer_name"],
            place=data["place"],
            address=data.get("address") or "",
            phone=data["phone"],
            dta_list=dta_list,
            products_json=data["products_json"],
            price=float(data["price"]),
            delivery_by=data["delivery_by"],
            payment_mode=data["payment_mode"],
            status="Delivered",
            delivery_type=data.get("delivery_type") or "Product Delivery",
            exch_old_dta=data.get("exch_old_dta"),
            exch_old_desc=data.get("exch_old_desc"),
            exch_old_value=float(data.get("exch_old_value") or 0.0),
            warranty_action=data.get("warranty_action"),
            jenny=int(data.get("jenny") or 0)
        )

        delivery_id = res.get("id")
        if delivery_id:
            # Auto-create corresponding Bill directly
            del_type = data.get("delivery_type") or "Product Delivery"
            bill_payload = {
                'date': date_val,
                'customer_name': data["customer_name"],
                'price': float(data["price"]),
                'payment_mode': data["payment_mode"],
                'note': f"Delivery #{delivery_id}",
                'delivery': 1,
                'jenny': int(data.get("jenny") or 0),
                'platform': data.get("platform") or "Regular Customer",
                'brand': '',
                'model': '',
                'dta': dta_list,
                'transaction_type': 'Sale',
                'products_json': data["products_json"],
                'exch_old_brand': '',
                'exch_old_model': data.get("exch_old_desc") or '',
                'exch_old_dta': data.get("exch_old_dta") or '',
                'exch_old_price': float(data.get("exch_old_value") or 0.0),
                'exch_balance': float(data["price"]) if del_type == 'Exchange' else 0.0,
                'exch_new_brand': '',
                'exch_new_model': '',
                'exch_new_dta': '',
                'exch_new_price': 0.0,
                'mixed_cash': 0.0,
                'mixed_card': 0.0,
                'mixed_tabby': 0.0,
                'mixed_tamara': 0.0,
                'mixed_bank': 0.0,
                'source': 'Inventory',
                'exch_old_source': 'Inventory'
            }

            try:
                p_list = json.loads(data["products_json"])
                if p_list and len(p_list) > 0:
                    bill_payload['brand'] = p_list[0].get('brand', '')
                    bill_payload['model'] = p_list[0].get('model', '')
                    bill_payload['dta'] = p_list[0].get('dta', '')
            except Exception:
                pass

            if del_type == 'Exchange':
                bill_payload['transaction_type'] = 'Exchange'
                try:
                    p_list = json.loads(data["products_json"])
                    if p_list and len(p_list) > 0:
                        bill_payload['exch_new_brand'] = p_list[0].get('brand', '')
                        bill_payload['exch_new_model'] = p_list[0].get('model', '')
                        bill_payload['exch_new_dta'] = p_list[0].get('dta', '')
                        bill_payload['exch_new_price'] = float(p_list[0].get('price', 0))
                except Exception:
                    pass
            elif del_type == 'Warranty (Return)':
                bill_payload['transaction_type'] = 'Return'

            new_bill = db.add_bill(bill_payload)
            if new_bill and new_bill.get("id"):
                db.update_delivery(delivery_id, {"bill_id": new_bill["id"], "status": "Delivered"})

        return jsonify(res)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/whatsapp-catalog-text", methods=["GET"])
def api_get_whatsapp_catalog_text():
    try:
        conn = sqlite3.connect(db.DB_PATH)
        cursor = conn.cursor()
        cursor.execute("CREATE TABLE IF NOT EXISTS system_settings (key TEXT PRIMARY KEY, value TEXT, updated_at DATETIME DEFAULT CURRENT_TIMESTAMP)")
        cursor.execute("SELECT value FROM system_settings WHERE key = 'whatsapp_catalog_text'")
        row = cursor.fetchone()
        conn.close()
        text_val = row[0] if row and row[0] else ""
        return jsonify({"raw_text": text_val})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/whatsapp-catalog-text", methods=["POST"])
def api_save_whatsapp_catalog_text():
    try:
        data = request.get_json() or {}
        raw_text = data.get("raw_text", "")
        conn = sqlite3.connect(db.DB_PATH)
        cursor = conn.cursor()
        cursor.execute("CREATE TABLE IF NOT EXISTS system_settings (key TEXT PRIMARY KEY, value TEXT, updated_at DATETIME DEFAULT CURRENT_TIMESTAMP)")
        cursor.execute("INSERT OR REPLACE INTO system_settings (key, value, updated_at) VALUES ('whatsapp_catalog_text', ?, CURRENT_TIMESTAMP)", (raw_text,))
        conn.commit()
        conn.close()
        return jsonify({"status": "ok", "raw_text": raw_text})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── CUSTOMER CRM ENDPOINTS ──────────────────────────────────────────────────

@app.route("/api/customers", methods=["GET"])
def api_customers_list():
    try:
        search = request.args.get("search")
        customers = db.get_all_customers(search=search)
        return jsonify(customers)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/customers", methods=["POST"])
def api_customers_create():
    try:
        data = request.get_json() or {}
        if not data.get("customer_name") or not data.get("mobile_number"):
            return jsonify({"error": "Customer Name and Mobile Number are required"}), 400
        res = db.add_customer(data)
        return jsonify(res), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/customers/<int:customer_id>", methods=["PUT"])
def api_customers_update(customer_id):
    try:
        data = request.get_json() or {}
        res = db.update_customer(customer_id, data)
        return jsonify(res)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/customers/<int:customer_id>", methods=["DELETE"])
def api_customers_delete(customer_id):
    try:
        res = db.delete_customer(customer_id)
        return jsonify(res)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/customers/import-sales", methods=["POST"])
def api_customers_import_sales():
    try:
        res = db.import_customers_from_sales_and_deliveries()
        return jsonify(res)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/deliveries/<int:delivery_id>", methods=["PUT"])
@permission_required("deliveries")
def api_deliveries_update(delivery_id):
    data = request.get_json()
    if not data:
        return jsonify({"error": "No data provided"}), 400
        
    try:
        # Fetch current delivery details to check if status is transitioning to Refunded
        delivery = db.get_delivery(delivery_id)
        if not delivery:
            return jsonify({"error": "Delivery not found"}), 404
            
        old_status = delivery.get("status")
        
        # Check if this is a full update vs status-only update
        if "customer_name" in data:
            # Full update flow
            required = ["customer_name", "place", "phone", "products_json", "price", "delivery_by", "payment_mode"]
            for f in required:
                val = data.get(f)
                if val is None or (isinstance(val, str) and not val.strip()):
                    return jsonify({"error": f"Field '{f}' is required"}), 400
            
            # Extract DTA list automatically from products_json
            import json
            try:
                p_list = json.loads(data["products_json"])
                dtas = [p.get("dta", "").strip().upper() for p in p_list if p.get("dta")]
                dta_list = ", ".join(dtas)
            except Exception:
                dta_list = ""
                
            new_status = data.get("status") or delivery.get("status") or "Billed Pending"
            
            res = db.update_delivery(
                delivery_id=delivery_id,
                data={
                    "date": data.get("date") or delivery.get("date"),
                    "customer_name": data["customer_name"],
                    "place": data["place"],
                    "address": data.get("address") or "",
                    "phone": data["phone"],
                    "dta_list": dta_list,
                    "products_json": data["products_json"],
                    "price": float(data["price"]),
                    "delivery_by": data["delivery_by"],
                    "payment_mode": data["payment_mode"],
                    "status": new_status,
                    "delivery_type": data.get("delivery_type") or "Product Delivery",
                    "exch_old_dta": data.get("exch_old_dta"),
                    "exch_old_desc": data.get("exch_old_desc"),
                    "exch_old_value": float(data.get("exch_old_value") or 0.0),
                    "warranty_action": data.get("warranty_action"),
                    "jenny": int(data.get("jenny") or 0),
                    "bill_id": data.get("bill_id")
                }
            )
        else:
            # Status-only update flow
            if "status" not in data:
                return jsonify({"error": "Field 'status' is required"}), 400
            new_status = data["status"]
            res = db.update_delivery_status(delivery_id, new_status)
        
        # If transitioning to Refunded, automatically create a Return bill
        if new_status == "Refunded" and old_status != "Refunded":
            # Check if a Return bill already exists for this delivery to prevent duplicates
            conn = db.get_connection()
            existing_return = conn.execute(
                "SELECT id FROM bills WHERE transaction_type = 'Return' AND note LIKE ?", 
                (f"%Auto-created from Refunded Delivery #{delivery_id}%",)
            ).fetchone()
            conn.close()
            
            if not existing_return:
                # Create the return bill
                bill_data = {
                    "date": today_str(), # Return happens today
                    "customer_name": data.get("customer_name") or delivery["customer_name"],
                    "price": float(data.get("price") or delivery["price"] or 0.0),
                    "payment_mode": data.get("payment_mode") or delivery["payment_mode"],
                    "transaction_type": "Return",
                    "platform": "Regular Customer", # default
                    "delivery": 0,
                    "products_json": data.get("products_json") or delivery["products_json"],
                    "note": f"Auto-created from Refunded Delivery #{delivery_id}",
                    "source": "Inventory" # default
                }
                db.create_bill(bill_data)
                
        return jsonify(res)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/deliveries/<int:delivery_id>", methods=["DELETE"])
@permission_required("deliveries")
def api_deliveries_delete(delivery_id):
    try:
        res = db.delete_delivery(delivery_id)
        return jsonify(res)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/deliveries/export-pdf", methods=["GET"])
@permission_required("deliveries")
def api_deliveries_export_pdf():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    except ImportError:
        return jsonify({"error": "reportlab not installed. Run: py -m pip install reportlab"}), 500

    date = request.args.get("date")
    month = request.args.get("month")  # MM-YYYY
    if month:
        deliveries = db.get_all_deliveries(month=month)
        # Format month label for display (e.g. "06-2025" -> "June 2025")
        try:
            from datetime import datetime as _dt2
            date_label = _dt2.strptime(month, "%m-%Y").strftime("%B %Y")
        except Exception:
            date_label = month
    else:
        if not date:
            date = today_str()
        deliveries = db.get_all_deliveries(date=date)
        date_label = date

    # Safe date parsing helper for chronological sorting
    def parse_date_safe(d):
        d_str = d.get("date") or ""
        try:
            from datetime import datetime as _dt
            return _dt.strptime(d_str.strip(), "%d-%m-%Y")
        except Exception:
            from datetime import datetime as _dt
            return _dt.min

    # Sort chronologically (oldest first, then oldest ID first)
    deliveries.sort(key=lambda x: (parse_date_safe(x), x.get("id") or 0))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    BRAND_COLOR = colors.HexColor("#4F46E5")
    DARK = colors.HexColor("#1F2937")
    LIGHT_GRAY = colors.HexColor("#F9FAFC")
    ALT_ROW = colors.HexColor("#F8FAFC")
    WHITE = colors.white
    MUTED = colors.HexColor("#6B7280")
    RED_C   = colors.HexColor("#EF4444")
    RED_BG  = colors.HexColor("#FEF2F2")
    GREEN   = colors.HexColor("#10B981")
    GREEN_BG = colors.HexColor("#ECFDF5")

    title_style = ParagraphStyle("title", fontName="Helvetica-Bold", fontSize=14, textColor=DARK, alignment=TA_LEFT, leading=16)
    sub_style = ParagraphStyle("sub", fontName="Helvetica", fontSize=8.5, textColor=MUTED, alignment=TA_LEFT)
    sub_right_style = ParagraphStyle("sub_right", parent=sub_style, alignment=TA_RIGHT, leading=11)
    
    th_style = ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, textColor=WHITE, alignment=TA_CENTER)
    th_left = ParagraphStyle("thl", fontName="Helvetica-Bold", fontSize=8, textColor=WHITE, alignment=TA_LEFT)
    
    td = ParagraphStyle("td", fontName="Helvetica", fontSize=7.5, textColor=DARK, alignment=TA_LEFT, leading=9)
    td_bold = ParagraphStyle("tdb", fontName="Helvetica-Bold", fontSize=7.5, textColor=BRAND_COLOR, alignment=TA_LEFT)
    tdc = ParagraphStyle("tdc", fontName="Helvetica", fontSize=7.5, textColor=DARK, alignment=TA_CENTER)
    tdr = ParagraphStyle("tdr", fontName="Helvetica-Bold", fontSize=7.5, textColor=DARK, alignment=TA_RIGHT)

    # Status badge pill helper
    def make_status_badge(status_text):
        text_color = colors.HexColor("#1F2937")
        bg_color = colors.HexColor("#F3F4F6")
        
        if status_text == "Delivered":
            bg_color = colors.HexColor("#D1FAE5")
            text_color = colors.HexColor("#065F46")
        elif status_text == "Out for Delivery":
            bg_color = colors.HexColor("#DBEAFE")
            text_color = colors.HexColor("#1E40AF")
        elif status_text in ("Pending", "Billed Pending"):
            bg_color = colors.HexColor("#FEF3C7")
            text_color = colors.HexColor("#92400E")
        elif status_text in ("Cancelled", "Refunded"):
            bg_color = colors.HexColor("#FEE2E2")
            text_color = colors.HexColor("#991B1B")

        badge_style = ParagraphStyle(
            f"badge_{status_text.replace(' ', '_')}",
            fontName="Helvetica-Bold",
            fontSize=6.5,
            textColor=text_color,
            alignment=TA_CENTER,
            leading=7
        )
        badge_p = Paragraph(status_text, badge_style)
        badge_table = Table([[badge_p]], colWidths=[14 * mm])
        badge_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), bg_color),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("LEFTPADDING", (0, 0), (-1, -1), 1),
            ("RIGHTPADDING", (0, 0), (-1, -1), 1),
        ]))
        return badge_table

    story = []

    # ── Header Banner (Clean Minimalist) ──────────────────────────────────────
    logo_path = os.path.join(os.path.dirname(__file__), "static", "logo.png")
    logo_widget = None
    if os.path.exists(logo_path):
        try:
            from reportlab.platypus import Image as RLImage
            logo_widget = RLImage(logo_path, width=28 * mm, height=6 * mm)
        except Exception:
            pass

    header_left_cells = []
    if logo_widget:
        header_left_cells.append(logo_widget)
        header_left_cells.append(Spacer(1, 1.5 * mm))
    else:
        logo_text_style = ParagraphStyle("logo_text", fontName="Helvetica-Bold", fontSize=14, textColor=BRAND_COLOR, alignment=TA_LEFT)
        header_left_cells.append(Paragraph("BUYOLOGY", logo_text_style))
        header_left_cells.append(Spacer(1, 1 * mm))
        
    header_left_cells.append(Paragraph("Deliveries Registry Report", title_style))

    from datetime import datetime
    right_text = f"<b>Period:</b> {date_label}<br/><b>Generated:</b> {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    header_right_p = Paragraph(right_text, sub_right_style)

    header_data = [[header_left_cells, header_right_p]]
    header_table = Table(header_data, colWidths=[120 * mm, 60 * mm])
    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 0), (-1, -1), 1.5, BRAND_COLOR),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 6 * mm))

    # Double-Row Summary Cards
    def is_jenny(d):
        val = d.get("jenny")
        return val is True or val == 1 or str(val).strip().lower() in ("true", "1")

    # Filter out jenny to get accurate statistics
    filtered_deliveries = [d for d in deliveries if not is_jenny(d)]
    
    # Row 1 Calculations
    net_rev = sum(
        (d.get("price") or 0.0) for d in filtered_deliveries
        if d.get("status") not in ("Cancelled", "Refunded")
        and (d.get("delivery_type") or "Product Delivery") != "Refund"
    )
    
    courier_rev = sum(
        (d.get("price") or 0.0) for d in filtered_deliveries
        if d.get("delivery_by") == "Courier"
        and d.get("status") not in ("Cancelled", "Refunded")
        and (d.get("delivery_type") or "Product Delivery") != "Refund"
    )
    courier_count = sum(1 for d in filtered_deliveries if d.get("delivery_by") == "Courier")
    
    guy_rev = sum(
        (d.get("price") or 0.0) for d in filtered_deliveries
        if d.get("delivery_by") == "Delivery Guy"
        and d.get("status") not in ("Cancelled", "Refunded")
        and (d.get("delivery_type") or "Product Delivery") != "Refund"
    )
    guy_count = sum(1 for d in filtered_deliveries if d.get("delivery_by") == "Delivery Guy")
    
    # Row 2 Calculations
    refund_count = sum(
        1 for d in filtered_deliveries
        if d.get("status") == "Refunded" or (d.get("delivery_type") or "Product Delivery") == "Refund"
    )
    refund_val = sum(
        (d.get("price") or 0.0) for d in filtered_deliveries
        if d.get("status") == "Refunded" or (d.get("delivery_type") or "Product Delivery") == "Refund"
    )
    
    warranty_return_count = sum(
        1 for d in filtered_deliveries
        if (d.get("delivery_type") or "Product Delivery") == "Warranty (Return)"
    )
    
    warranty_delivery_count = sum(
        1 for d in filtered_deliveries
        if (d.get("delivery_type") or "Product Delivery") == "Warranty Delivery"
    )

    def make_card(value_html, label_html, subtext_html, bg_color, border_color, text_color):
        card_data = [
            [Paragraph(value_html, ParagraphStyle("card_val", fontName="Helvetica-Bold", fontSize=11, textColor=text_color, alignment=TA_CENTER, leading=13))],
            [Paragraph(label_html, ParagraphStyle("card_lbl", fontName="Helvetica-Bold", fontSize=7.5, textColor=colors.HexColor("#374151"), alignment=TA_CENTER, leading=9))],
            [Paragraph(subtext_html, ParagraphStyle("card_sub", fontName="Helvetica", fontSize=6.5, textColor=colors.HexColor("#6B7280"), alignment=TA_CENTER, leading=8))]
        ]
        card_table = Table(card_data, colWidths=[56 * mm])
        card_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), bg_color),
            ("BOX", (0, 0), (-1, -1), 0.5, border_color),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ]))
        return card_table

    # Row 1 cards
    row1_data = [[
        make_card(f"AED {net_rev:,.2f}", "Net Delivery Revenue", "Excl. Cancelled & Refunds", colors.HexColor("#EEF2FF"), colors.HexColor("#C7D2FE"), colors.HexColor("#4F46E5")),
        make_card(f"AED {courier_rev:,.2f}", "Courier Revenue", f"{courier_count} Shipments via Courier", colors.HexColor("#F8FAFC"), colors.HexColor("#E2E8F0"), colors.HexColor("#1F2937")),
        make_card(f"AED {guy_rev:,.2f}", "Delivery Guy Revenue", f"{guy_count} Shipments in-house", colors.HexColor("#F8FAFC"), colors.HexColor("#E2E8F0"), colors.HexColor("#1F2937"))
    ]]
    row1_table = Table(row1_data, colWidths=[60 * mm, 60 * mm, 60 * mm])
    row1_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    
    # Row 2 cards
    row2_data = [[
        make_card(f"−AED {refund_val:,.2f}" if refund_val > 0 else f"AED {refund_val:,.2f}", "Total Refunds", f"{refund_count} Refunds processed", colors.HexColor("#FEF2F2"), colors.HexColor("#FCA5A5"), colors.HexColor("#EF4444")),
        make_card(f"{warranty_return_count} Returns", "Warranty Collections", "Items picked up from customer", colors.HexColor("#FFF7ED"), colors.HexColor("#FED7AA"), colors.HexColor("#D97706")),
        make_card(f"{warranty_delivery_count} Deliveries", "Warranty Dispatches", "Repaired/replaced items sent", colors.HexColor("#ECFDF5"), colors.HexColor("#A7F3D0"), colors.HexColor("#059669"))
    ]]
    row2_table = Table(row2_data, colWidths=[60 * mm, 60 * mm, 60 * mm])
    row2_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    story.append(row1_table)
    story.append(Spacer(1, 4 * mm))
    story.append(row2_table)
    story.append(Spacer(1, 6 * mm))

    # ── Separation of Tables ──────────────────────────────────────────────────
    completed_deliveries = [d for d in deliveries if d.get("status") == "Delivered"]
    other_deliveries = [d for d in deliveries if d.get("status") != "Delivered"]

    section_title_style = ParagraphStyle(
        "sectitle",
        fontName="Helvetica-Bold",
        fontSize=10,
        textColor=BRAND_COLOR,
        alignment=TA_LEFT,
        spaceAfter=3 * mm
    )

    # ── Table 1: Completed Deliveries (Delivered Only) ──
    story.append(Paragraph("<b>Completed Deliveries</b>", section_title_style))

    completed_headers = [
        Paragraph("<b>#</b>", th_style),
        Paragraph("<b>Date</b>", th_style),
        Paragraph("<b>Customer & Address</b>", th_left),
        Paragraph("<b>Product Details</b>", th_left),
        Paragraph("<b>Total Price</b>", th_style),
        Paragraph("<b>Logistics</b>", th_left),
    ]
    completed_table_data = [completed_headers]

    for idx, d in enumerate(completed_deliveries):
        d_id = d.get("id")
        d_date = d.get("date")
        
        name = d.get("customer_name")
        is_jenny = int(d.get("jenny") or 0) == 1
        dtype = d.get("delivery_type") or "Product Delivery"
        
        badges = []
        if dtype == "Exchange":
            badges.append("<font color='#EC4899'><b>Exchange</b></font>")
        elif dtype == "Warranty (Return)":
            badges.append("<font color='#F59E0B'><b>Warranty (Return)</b></font>")
        elif dtype == "Warranty Delivery":
            badges.append("<font color='#F59E0B'><b>Warranty Delivery</b></font>")
        elif dtype == "Refund":
            badges.append("<font color='#EF4444'><b>Refund</b></font>")
        
        if is_jenny:
            badges.append("<font color='#be185d'><b>Jenny</b></font>")
            
        name_text = f"<b>{name}</b>"
        if badges:
            name_text += f" ({', '.join(badges)})"

        phone = d.get("phone")
        place = d.get("place")
        address = d.get("address") or ""
        
        addr_text = f"{name_text}<br/>Ph: {phone}<br/>Place: {place}"
        if address:
            addr_text += f"<br/>Addr: {address}"
        customer_p = Paragraph(addr_text, td)

        import json
        p_list = []
        try:
            p_list = json.loads(d.get("products_json") or "[]")
        except Exception:
            pass
        
        prod_lines = []
        for p_idx, p in enumerate(p_list):
            brand = p.get("brand", "").strip()
            model = p.get("model", "").split("|")[0].strip()
            qty = p.get("quantity", 1) or 1
            price = p.get("price", 0.0)
            p_dta = p.get("dta", "")
            prod_lines.append(f"• <b>{brand} {model}</b> (x{qty}) [DTA: {p_dta}] (AED {price:,.2f} each)")
        
        if dtype == "Exchange":
            old_dta = d.get("exch_old_dta") or ""
            old_desc = d.get("exch_old_desc") or ""
            old_val = d.get("exch_old_value") or 0.0
            prod_lines.append(f"<font color='#EC4899'><b>[Exchange Return]</b> {old_desc} [DTA: {old_dta}] (Value: AED {old_val:,.2f})</font>")
        elif dtype == "Warranty (Return)":
            waction = d.get("warranty_action") or "Pickup/Return"
            prod_lines.append(f"<font color='#F59E0B'><b>[Warranty Issue]</b> {waction}</font>")

        prod_text = "<br/>".join(prod_lines) if prod_lines else "—"
        product_p = Paragraph(prod_text, td)

        price_raw = d.get('price', 0.0) or 0.0
        status_val = d.get("status", "Billed Pending")

        tdr_red = ParagraphStyle("tdrr", fontName="Helvetica-Bold", fontSize=7.5, textColor=RED_C, alignment=TA_RIGHT)
        tdr_mut = ParagraphStyle("tdrm", fontName="Helvetica", fontSize=7.5, textColor=MUTED, alignment=TA_RIGHT)
        stat_sub_del = ParagraphStyle("ssldel", fontName="Helvetica-Oblique", fontSize=6.5, textColor=MUTED, alignment=TA_RIGHT)

        if status_val == "Refunded" or dtype == "Refund":
            price_p = [
                Paragraph(f"<b>−AED {price_raw:,.2f}</b>", tdr_red),
                Paragraph("↩ Refunded" if status_val == "Refunded" else "↩ Refund", stat_sub_del),
            ]
        elif is_jenny:
            price_p = [
                Paragraph(f"AED {price_raw:,.2f}", tdr_mut),
                Paragraph("(Jenny)", stat_sub_del),
            ]
        else:
            price_p = Paragraph(f"AED {price_raw:,.2f}", tdr)

        logistic_text = f"<b>By:</b> {d.get('delivery_by')}<br/><b>MOP:</b> {d.get('payment_mode')}"
        logistic_p = Paragraph(logistic_text, td)

        completed_table_data.append([
            Paragraph(f"{idx + 1}<br/><font color='#6c757d'>#{d_id}</font>", tdc),
            Paragraph(d_date, tdc),
            customer_p,
            product_p,
            price_p,
            logistic_p,
        ])

    if not completed_deliveries:
        completed_table_data.append([
            Paragraph("<font color='#6B7280'>No completed deliveries found in this period.</font>", td),
            "", "", "", "", ""
        ])

    completed_col_widths = [12 * mm, 17 * mm, 45 * mm, 62 * mm, 18 * mm, 26 * mm]
    completed_table_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_COLOR),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, 0), 1.5, colors.HexColor("#312E81")),
        ("LINEBELOW", (0, 1), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
    ])

    if not completed_deliveries:
        completed_table_style.add("SPAN", (0, 1), (-1, 1))
        completed_table_style.add("ALIGN", (0, 1), (-1, 1), "CENTER")
    else:
        for i in range(1, len(completed_table_data)):
            d = completed_deliveries[i - 1]
            dtype = d.get("delivery_type") or "Product Delivery"
            if d.get("status") == "Refunded" or dtype == "Refund":
                completed_table_style.add("BACKGROUND", (0, i), (-1, i), RED_BG)
            elif i % 2 == 1:
                completed_table_style.add("BACKGROUND", (0, i), (-1, i), ALT_ROW)
            else:
                completed_table_style.add("BACKGROUND", (0, i), (-1, i), WHITE)

    comp_t = Table(completed_table_data, colWidths=completed_col_widths)
    comp_t.setStyle(completed_table_style)
    story.append(comp_t)
    story.append(Spacer(1, 6 * mm))

    # ── Table 2: Active & Exception Deliveries ──
    story.append(Paragraph("<b>Active & Exception Deliveries</b>", section_title_style))

    active_headers = [
        Paragraph("<b>#</b>", th_style),
        Paragraph("<b>Date</b>", th_style),
        Paragraph("<b>Customer & Address</b>", th_left),
        Paragraph("<b>Product Details</b>", th_left),
        Paragraph("<b>Total Price</b>", th_style),
        Paragraph("<b>Logistics</b>", th_left),
        Paragraph("<b>Status</b>", th_style),
    ]
    active_table_data = [active_headers]

    for idx, d in enumerate(other_deliveries):
        d_id = d.get("id")
        d_date = d.get("date")
        
        name = d.get("customer_name")
        is_jenny = int(d.get("jenny") or 0) == 1
        dtype = d.get("delivery_type") or "Product Delivery"
        
        badges = []
        if dtype == "Exchange":
            badges.append("<font color='#EC4899'><b>Exchange</b></font>")
        elif dtype == "Warranty (Return)":
            badges.append("<font color='#F59E0B'><b>Warranty (Return)</b></font>")
        elif dtype == "Warranty Delivery":
            badges.append("<font color='#F59E0B'><b>Warranty Delivery</b></font>")
        elif dtype == "Refund":
            badges.append("<font color='#EF4444'><b>Refund</b></font>")
        
        if is_jenny:
            badges.append("<font color='#be185d'><b>Jenny</b></font>")
            
        name_text = f"<b>{name}</b>"
        if badges:
            name_text += f" ({', '.join(badges)})"

        phone = d.get("phone")
        place = d.get("place")
        address = d.get("address") or ""
        
        addr_text = f"{name_text}<br/>Ph: {phone}<br/>Place: {place}"
        if address:
            addr_text += f"<br/>Addr: {address}"
        customer_p = Paragraph(addr_text, td)

        import json
        p_list = []
        try:
            p_list = json.loads(d.get("products_json") or "[]")
        except Exception:
            pass
        
        prod_lines = []
        for p_idx, p in enumerate(p_list):
            brand = p.get("brand", "").strip()
            model = p.get("model", "").split("|")[0].strip()
            qty = p.get("quantity", 1) or 1
            price = p.get("price", 0.0)
            p_dta = p.get("dta", "")
            prod_lines.append(f"• <b>{brand} {model}</b> (x{qty}) [DTA: {p_dta}] (AED {price:,.2f} each)")
        
        if dtype == "Exchange":
            old_dta = d.get("exch_old_dta") or ""
            old_desc = d.get("exch_old_desc") or ""
            old_val = d.get("exch_old_value") or 0.0
            prod_lines.append(f"<font color='#EC4899'><b>[Exchange Return]</b> {old_desc} [DTA: {old_dta}] (Value: AED {old_val:,.2f})</font>")
        elif dtype == "Warranty (Return)":
            waction = d.get("warranty_action") or "Pickup/Return"
            prod_lines.append(f"<font color='#F59E0B'><b>[Warranty Issue]</b> {waction}</font>")

        prod_text = "<br/>".join(prod_lines) if prod_lines else "—"
        product_p = Paragraph(prod_text, td)

        price_raw = d.get('price', 0.0) or 0.0
        status_val = d.get("status", "Billed Pending")

        tdr_red = ParagraphStyle("tdrr", fontName="Helvetica-Bold", fontSize=7.5, textColor=RED_C, alignment=TA_RIGHT)
        tdr_mut = ParagraphStyle("tdrm", fontName="Helvetica", fontSize=7.5, textColor=MUTED, alignment=TA_RIGHT)
        stat_sub_del = ParagraphStyle("ssldel", fontName="Helvetica-Oblique", fontSize=6.5, textColor=MUTED, alignment=TA_RIGHT)

        if status_val == "Refunded" or dtype == "Refund":
            price_p = [
                Paragraph(f"<b>−AED {price_raw:,.2f}</b>", tdr_red),
                Paragraph("↩ Refunded" if status_val == "Refunded" else "↩ Refund", stat_sub_del),
            ]
        elif is_jenny:
            price_p = [
                Paragraph(f"AED {price_raw:,.2f}", tdr_mut),
                Paragraph("(Jenny)", stat_sub_del),
            ]
        else:
            price_p = Paragraph(f"AED {price_raw:,.2f}", tdr)

        logistic_text = f"<b>By:</b> {d.get('delivery_by')}<br/><b>MOP:</b> {d.get('payment_mode')}"
        logistic_p = Paragraph(logistic_text, td)

        status_badge = make_status_badge(status_val)

        active_table_data.append([
            Paragraph(f"{idx + 1}<br/><font color='#6c757d'>#{d_id}</font>", tdc),
            Paragraph(d_date, tdc),
            customer_p,
            product_p,
            price_p,
            logistic_p,
            status_badge,
        ])

    if not other_deliveries:
        active_table_data.append([
            Paragraph("<font color='#6B7280'>No active or exception deliveries found in this period.</font>", td),
            "", "", "", "", "", ""
        ])

    active_col_widths = [12 * mm, 17 * mm, 40 * mm, 52 * mm, 18 * mm, 26 * mm, 15 * mm]
    active_table_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_COLOR),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, 0), 1.5, colors.HexColor("#312E81")),
        ("LINEBELOW", (0, 1), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
    ])

    if not other_deliveries:
        active_table_style.add("SPAN", (0, 1), (-1, 1))
        active_table_style.add("ALIGN", (0, 1), (-1, 1), "CENTER")
    else:
        for i in range(1, len(active_table_data)):
            d = other_deliveries[i - 1]
            dtype = d.get("delivery_type") or "Product Delivery"
            if d.get("status") == "Refunded" or dtype == "Refund":
                active_table_style.add("BACKGROUND", (0, i), (-1, i), RED_BG)
            elif i % 2 == 1:
                active_table_style.add("BACKGROUND", (0, i), (-1, i), ALT_ROW)
            else:
                active_table_style.add("BACKGROUND", (0, i), (-1, i), WHITE)

    act_t = Table(active_table_data, colWidths=active_col_widths)
    act_t.setStyle(active_table_style)
    story.append(act_t)

    # ── Table 3: Performance Summary by Delivery Type ──
    story.append(Spacer(1, 8 * mm))
    story.append(Paragraph("<b>Delivery Type Performance Summary</b>", section_title_style))
    
    summary_headers = [
        Paragraph("<b>Delivery Type</b>", th_left),
        Paragraph("<b>Total Count</b>", th_style),
        Paragraph("<b>Delivered</b>", th_style),
        Paragraph("<b>Billed Pending</b>", th_style),
        Paragraph("<b>Pending</b>", th_style),
        Paragraph("<b>Out for Delivery</b>", th_style),
        Paragraph("<b>Cancelled</b>", th_style),
        Paragraph("<b>Total Price (AED)</b>", th_style)
    ]
    
    summary_table_data = [summary_headers]
    
    summary_types = [
        ("Product Delivery", "Product Delivery"),
        ("Exchange", "Total Exchanges"),
        ("Warranty (Return)", "Warranty Returns"),
        ("Warranty Delivery", "Warranty Deliveries"),
        ("Refund", "Total Refunds")
    ]
    
    td_left_bold = ParagraphStyle("tdlb", fontName="Helvetica-Bold", fontSize=7.5, textColor=DARK, alignment=TA_LEFT)
    tdc_bold = ParagraphStyle("tdcb", fontName="Helvetica-Bold", fontSize=7.5, textColor=DARK, alignment=TA_CENTER)
    tdr_bold_purple = ParagraphStyle("tdrbp", fontName="Helvetica-Bold", fontSize=7.5, textColor=BRAND_COLOR, alignment=TA_RIGHT)
    tdr_bold_red = ParagraphStyle("tdrbr", fontName="Helvetica-Bold", fontSize=7.5, textColor=RED_C, alignment=TA_RIGHT)
    
    for t_key, t_label in summary_types:
        type_list = [d for d in deliveries if (d.get("delivery_type") or "Product Delivery") == t_key]
        total_type = len(type_list)
        del_type = sum(1 for d in type_list if d.get("status") == "Delivered")
        bp_type = sum(1 for d in type_list if d.get("status") == "Billed Pending")
        p_type = sum(1 for d in type_list if d.get("status") == "Pending")
        out_type = sum(1 for d in type_list if d.get("status") == "Out for Delivery")
        can_type = sum(1 for d in type_list if d.get("status") == "Cancelled")
        
        rev_type = sum(
            (d.get("price") or 0.0) for d in type_list
            if d.get("status") not in ("Cancelled", "Refunded") and int(d.get("jenny") or 0) == 0
        )
        
        sign = -1.0 if t_key in ("Refund", "Refunded") else 1.0
        display_rev = rev_type * sign
        
        rev_style = tdr_bold_red if display_rev < 0 else tdr_bold_purple
        
        summary_table_data.append([
            Paragraph(t_label, td_left_bold),
            Paragraph(str(total_type), tdc_bold),
            Paragraph(str(del_type), tdc),
            Paragraph(str(bp_type), tdc),
            Paragraph(str(p_type), tdc),
            Paragraph(str(out_type), tdc),
            Paragraph(str(can_type), tdc),
            Paragraph(f"AED {display_rev:,.2f}" if display_rev >= 0 else f"−AED {abs(display_rev):,.2f}", rev_style)
        ])
        
    summary_col_widths = [45 * mm, 18 * mm, 18 * mm, 20 * mm, 18 * mm, 23 * mm, 18 * mm, 20 * mm]
    summary_table_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_COLOR),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, 0), 1.5, colors.HexColor("#312E81")),
        ("LINEBELOW", (0, 1), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
    ])
    
    for i in range(1, len(summary_table_data)):
        if i % 2 == 1:
            summary_table_style.add("BACKGROUND", (0, i), (-1, i), ALT_ROW)
        else:
            summary_table_style.add("BACKGROUND", (0, i), (-1, i), WHITE)
            
    sum_t = Table(summary_table_data, colWidths=summary_col_widths)
    sum_t.setStyle(summary_table_style)
    story.append(sum_t)

    doc.build(story)
    buf.seek(0)

    filename = f"Deliveries_Report_{_date.today().strftime('%Y-%m-%d')}.pdf"
    response = send_file(
        buf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename
    )
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.route("/api/deliveries/export-excel", methods=["GET"])
@permission_required("deliveries")
def api_deliveries_export_excel():
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500

    date = request.args.get("date")
    month = request.args.get("month")  # MM-YYYY
    if month:
        deliveries = db.get_all_deliveries(month=month)
        try:
            from datetime import datetime as _dt2
            date_label = _dt2.strptime(month, "%m-%Y").strftime("%B %Y")
        except Exception:
            date_label = month
    else:
        if not date:
            date = today_str()
        deliveries = db.get_all_deliveries(date=date)
        date_label = date

    # Safe date parsing helper for chronological sorting
    def parse_date_safe(d):
        d_str = d.get("date") or ""
        try:
            from datetime import datetime as _dt
            return _dt.strptime(d_str.strip(), "%d-%m-%Y")
        except Exception:
            from datetime import datetime as _dt
            return _dt.min

    # Sort chronologically (oldest first, then oldest ID first)
    deliveries.sort(key=lambda x: (parse_date_safe(x), x.get("id") or 0))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Deliveries"

    # Theme colors
    BRAND_COLOR = "402F75"
    HEADER_FILL = PatternFill("solid", fgColor=BRAND_COLOR)
    WHITE_BOLD_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    HEADER_FONT = Font(bold=True, name="Calibri", size=13)
    BOLD_FONT = Font(bold=True, name="Calibri", size=11)
    REGULAR_FONT = Font(name="Calibri", size=11)
    ALT_ROW_FILL = PatternFill("solid", fgColor="F8F7FA")
    
    # Status styling
    STATUS_COLORS = {
        "Billed Pending": ("FFE6C7", "B35C00"),       # soft orange
        "Pending": ("FFE6C7", "B35C00"),              # soft orange
        "Out for Delivery": ("E3F2FD", "0D47A1"),  # soft blue
        "Delivered": ("E8F5E9", "1B5E20"),     # soft green
        "Cancelled": ("FFEBEE", "B71C1C"),     # soft red
        "Refunded": ("FFF0F0", "DC3545")       # soft red/pink
    }

    thin_border = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3")
    )

    # Title Banner
    ws.merge_cells("A1:R1")
    title_cell = ws["A1"]
    title_cell.value = f"Buyology Deliveries Report — {date_label}"
    title_cell.fill = PatternFill("solid", fgColor="FCBC12") # accent yellow
    title_cell.font = HEADER_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Headers
    headers = [
        "ID", "Date", "Customer Name", "Place", "Address", "Phone",
        "DTA(s)", "Product Details", "Total Price (AED)", "Delivery By", "Payment Mode", "Status",
        "Delivery Type", "Exchanged Item (Old)", "Returned Item Value (AED)", "Warranty Issue", "Jenny"
    ]
    ws.append([]) # row 2 empty spacer
    
    # Add headers to row 3
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[3].height = 24

    # Rows
    for row_num, d in enumerate(deliveries, 4):
        # Format product details
        import json
        p_list = []
        try:
            p_list = json.loads(d.get("products_json") or "[]")
        except Exception:
            pass
        
        prod_strs = []
        for p in p_list:
            brand = p.get("brand", "").strip()
            model = p.get("model", "").split("|")[0].strip()
            qty = p.get("quantity", 1) or 1
            price = p.get("price", 0.0)
            prod_strs.append(f"{brand} {model} (x{qty}) - AED {price:,.2f}")
        
        prod_details_str = ", ".join(prod_strs) if prod_strs else "—"

        row_data = [
            d.get("id"),
            d.get("date"),
            d.get("customer_name"),
            d.get("place"),
            d.get("address") or "—",
            d.get("phone"),
            d.get("dta_list"),
            prod_details_str,
            d.get("price"),
            d.get("delivery_by"),
            d.get("payment_mode"),
            d.get("status"),
            d.get("delivery_type") or "Product Delivery",
            f"{d.get('exch_old_dta') or ''} - {d.get('exch_old_desc') or ''}" if d.get("exch_old_dta") else "",
            d.get("exch_old_value") if d.get("exch_old_value") else "",
            d.get("warranty_action") or "",
            "Yes" if int(d.get("jenny") or 0) == 1 else "No"
        ]

        row_fill = ALT_ROW_FILL if row_num % 2 == 1 else None
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = thin_border
            cell.font = REGULAR_FONT
            
            # Alignments
            if col_idx in [1, 2, 6, 10, 11, 12, 17]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 9: # Price
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0.00"
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            if row_fill:
                cell.fill = row_fill
                
        # Style Status Cell
        status_val = d.get("status", "Billed Pending")
        if status_val in STATUS_COLORS:
            bg_color, fg_color = STATUS_COLORS[status_val]
            stat_cell = ws.cell(row=row_num, column=12)
            stat_cell.fill = PatternFill("solid", fgColor=bg_color)
            stat_cell.font = Font(color=fg_color, bold=True, name="Calibri", size=11)

    # Column widths
    widths = [6, 14, 22, 16, 26, 16, 16, 45, 18, 16, 18, 18, 18, 24, 24, 24, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"Deliveries_{date}.xlsx"
    response = send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response



# ── Run ───────────────────────────────────────────────────────────────────────


# ==========================================================
# WARRANTY ENDPOINTS
# ==========================================================

@app.route('/api/warranty', methods=['GET'])
def get_warranty_claims():
    try:
        claims = db.get_all_warranty_claims()
        return jsonify(claims)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/warranty', methods=['POST'])
def add_warranty_claim():
    try:
        data = request.json
        claim_id = db.create_warranty_claim(data)
        return jsonify({'success': True, 'id': claim_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/warranty/<int:claim_id>/resolve', methods=['PUT'])
def resolve_warranty_claim(claim_id):
    try:
        data = request.json
        db.resolve_warranty_claim(claim_id, data)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/warranty/<int:claim_id>/edit', methods=['POST'])
def edit_warranty_claim(claim_id):
    try:
        data = request.json
        db.edit_warranty_claim(claim_id, data)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/warranty/<int:claim_id>', methods=['DELETE'])
def delete_warranty_claim(claim_id):
    try:
        db.delete_warranty_claim(claim_id)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/warranty/lookup/<path:dta>', methods=['GET'])
def warranty_lookup(dta):
    dta_clean = dta.strip().upper()
    try:
        conn = db.get_connection()
        # Look in bills
        bill = conn.execute("SELECT brand, model, date, customer_name, price FROM bills WHERE dta = ? OR products_json LIKE ?", (dta_clean, f'%"{dta_clean}"%')).fetchone()
        if bill and bill['brand']:
            conn.close()
            return jsonify({
                'dta': dta_clean, 
                'brand': bill['brand'], 
                'model': bill['model'],
                'purchase_date': bill['date'],
                'customer_name': bill['customer_name'],
                'price': bill['price'],
                'source': 'sold'
            })
        
        # Look in inventory
        inv = conn.execute("SELECT brand, model, price, status FROM inventory_units WHERE dta = ?", (dta_clean,)).fetchone()
        if inv:
            conn.close()
            return jsonify({
                'dta': dta_clean, 
                'brand': inv['brand'], 
                'model': inv['model'],
                'price': inv['price'],
                'status': inv['status'],
                'source': 'inventory'
            })
            
        conn.close()
        return jsonify({'error': 'Not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route("/api/warranty/export-pdf", methods=["GET"])
def export_warranty_pdf():
    status_filter = request.args.get("status", "Pending")
    search_query = request.args.get("search", "").strip().lower()

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from datetime import datetime as datetime_cls
        from datetime import date as date_cls
        import json
    except ImportError:
        return jsonify({"error": "reportlab not installed. Run: py -m pip install reportlab"}), 500

    try:
        claims = db.get_all_warranty_claims()
        
        # 1. Filter by status (tab)
        if status_filter == "Pending":
            filtered = [c for c in claims if c.get("status") == "Pending"]
        elif status_filter == "Resolved":
            filtered = [c for c in claims if c.get("status") != "Pending"]
        else:
            filtered = claims

        # 2. Filter by search query
        if search_query:
            matching = []
            for c in filtered:
                cust = (c.get("customer_name") or "").lower()
                phone = (c.get("phone_number") or "").lower()
                loc = (c.get("location") or "").lower()
                note = (c.get("issue_note") or "").lower()
                r_note = (c.get("repair_note") or "").lower()
                dta = (c.get("dta") or "").lower()
                brand = (c.get("brand") or "").lower()
                model = (c.get("model") or "").lower()
                
                # Check items in products_json
                prod_match = False
                p_json = c.get("products_json")
                if p_json:
                    try:
                        p_list = json.loads(p_json)
                        for p in p_list:
                            p_dta = (p.get("dta") or "").lower()
                            p_brand = (p.get("brand") or "").lower()
                            p_model = (p.get("model") or "").lower()
                            if search_query in p_dta or search_query in p_brand or search_query in p_model:
                                prod_match = True
                                break
                    except Exception:
                        pass
                
                if (search_query in cust or 
                    search_query in phone or 
                    search_query in loc or 
                    search_query in note or 
                    search_query in r_note or 
                    search_query in dta or 
                    search_query in brand or 
                    search_query in model or 
                    prod_match):
                    matching.append(c)
            filtered = matching

        # Sort chronologically ascending (oldest first) — same order as the UI table
        filtered = sorted(filtered, key=lambda c: (c.get("claim_date") or "", c.get("id") or 0))

        # Calculate statistics
        total_claims = len(filtered)
        pending_count = sum(1 for c in filtered if c.get("status") == "Pending")
        resolved_count = total_claims - pending_count
        replaced_count = sum(1 for c in filtered if c.get("status") == "Replaced")
        refunded_count = sum(1 for c in filtered if c.get("status") in ("Returned Unfixed", "Refunded"))
        
        total_cost = 0.0
        for c in filtered:
            total_cost += float(c.get("repair_cost") or 0.0)
            total_cost += float(c.get("exch_balance") or 0.0)
            total_cost += float(c.get("refund_amount") or 0.0)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            leftMargin=15 * mm,
            rightMargin=15 * mm,
            topMargin=15 * mm,
            bottomMargin=15 * mm,
        )

        BRAND_COLOR = colors.HexColor("#402F75")  # Dark Purple
        DARK = colors.HexColor("#1F1B2C")
        LIGHT_GRAY = colors.HexColor("#F8F7FA")
        ALT_ROW = colors.HexColor("#F1EEF6")
        WHITE = colors.white
        MUTED = colors.HexColor("#6C7280")

        # Text styles
        title_style = ParagraphStyle("title", fontName="Helvetica-Bold", fontSize=14, textColor=DARK, alignment=TA_LEFT, leading=16)
        sub_style = ParagraphStyle("sub", fontName="Helvetica", fontSize=8.5, textColor=MUTED, alignment=TA_LEFT)
        sub_right_style = ParagraphStyle("sub_right", parent=sub_style, alignment=TA_RIGHT, leading=11)
        
        th_style = ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, textColor=WHITE, alignment=TA_CENTER)
        th_left = ParagraphStyle("thl", fontName="Helvetica-Bold", fontSize=8, textColor=WHITE, alignment=TA_LEFT)
        
        td = ParagraphStyle("td", fontName="Helvetica", fontSize=7.5, textColor=DARK, alignment=TA_LEFT, leading=9.5)
        tdc = ParagraphStyle("tdc", fontName="Helvetica", fontSize=7.5, textColor=DARK, alignment=TA_CENTER, leading=9.5)
        tdr = ParagraphStyle("tdr", fontName="Helvetica-Bold", fontSize=7.5, textColor=DARK, alignment=TA_RIGHT, leading=9.5)

        # Status badge helper
        def make_status_badge(status_text):
            text_color = colors.HexColor("#1F2937")
            bg_color = colors.HexColor("#F3F4F6")
            
            if status_text == "Pending":
                bg_color = colors.HexColor("#FEF3C7")
                text_color = colors.HexColor("#92400E")
            elif status_text == "Repaired":
                bg_color = colors.HexColor("#D1FAE5")
                text_color = colors.HexColor("#065F46")
            elif status_text == "Replaced":
                bg_color = colors.HexColor("#DBEAFE")
                text_color = colors.HexColor("#1E40AF")
            elif status_text in ("Returned Unfixed", "Returned", "Refunded"):
                bg_color = colors.HexColor("#FEE2E2")
                text_color = colors.HexColor("#991B1B")

            badge_style = ParagraphStyle(
                f"badge_{status_text.replace(' ', '_')}",
                fontName="Helvetica-Bold",
                fontSize=6.5,
                textColor=text_color,
                alignment=TA_CENTER,
                leading=7
            )
            badge_p = Paragraph(status_text, badge_style)
            badge_table = Table([[badge_p]], colWidths=[20 * mm])
            badge_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), bg_color),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 2.5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
                ("LEFTPADDING", (0, 0), (-1, -1), 1),
                ("RIGHTPADDING", (0, 0), (-1, -1), 1),
            ]))
            return badge_table

        # Card widget helper
        def make_card(value_html, label_html, subtext_html, bg_color, border_color, text_color):
            card_data = [
                [Paragraph(value_html, ParagraphStyle("card_val", fontName="Helvetica-Bold", fontSize=11, textColor=text_color, alignment=TA_CENTER, leading=13))],
                [Paragraph(label_html, ParagraphStyle("card_lbl", fontName="Helvetica-Bold", fontSize=7.5, textColor=colors.HexColor("#374151"), alignment=TA_CENTER, leading=9))],
                [Paragraph(subtext_html, ParagraphStyle("card_sub", fontName="Helvetica", fontSize=6.5, textColor=colors.HexColor("#6B7280"), alignment=TA_CENTER, leading=8))]
            ]
            card_table = Table(card_data, colWidths=[62 * mm])
            card_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), bg_color),
                ("BOX", (0, 0), (-1, -1), 0.5, border_color),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]))
            return card_table

        story = []

        # Branded header banner
        logo_path = os.path.join(os.path.dirname(__file__), "static", "logo.png")
        logo_widget = None
        if os.path.exists(logo_path):
            try:
                from reportlab.platypus import Image as RLImage
                logo_widget = RLImage(logo_path, width=42 * mm, height=9 * mm)
            except Exception:
                pass

        header_left = logo_widget if logo_widget else Paragraph("<b>BUYOLOGY</b>", ParagraphStyle("h_logo", fontName="Helvetica-Bold", fontSize=18, textColor=WHITE, alignment=TA_LEFT))

        header_title_style = ParagraphStyle(
            "h_title", fontName="Helvetica-Bold", fontSize=13,
            textColor=WHITE, alignment=TA_LEFT, leading=15
        )
        header_sub_style = ParagraphStyle(
            "h_sub", fontName="Helvetica-Bold", fontSize=9.5,
            textColor=colors.HexColor("#FCBC12"), alignment=TA_LEFT
        )
        header_sub_right_style = ParagraphStyle(
            "h_sub_right", parent=header_sub_style, alignment=TA_RIGHT, leading=11
        )

        from datetime import datetime as datetime_cls
        from datetime import date as date_cls
        right_text = f"<b>Queue:</b> {status_filter} Claims<br/><b>Generated:</b> {date_cls.today().strftime('%d-%m-%Y')} at {datetime_cls.now().strftime('%H:%M')}"
        if search_query:
            right_text += f"<br/><b>Search:</b> \"{search_query}\""
        header_right_p = Paragraph(right_text, header_sub_right_style)

        header_data = [[
            header_left,
            Paragraph("Warranty Claims Registry Report", header_title_style),
            header_right_p,
        ]]
        header_table = Table(header_data, colWidths=[80 * mm, 120 * mm, 67 * mm])
        header_table.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, -1), BRAND_COLOR),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING",  (0, 0), (-1, -1), 12),
            ("RIGHTPADDING", (0, 0), (-1, -1), 12),
            ("TOPPADDING",   (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
            ("ALIGN",        (2, 0), (2, 0),   "RIGHT"),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 5 * mm))

        # Statistics Cards Summary
        row1_data = [[
            make_card(f"{total_claims} Claims", "Total Claims", f"{pending_count} Pending | {resolved_count} Resolved", colors.HexColor("#EEF2FF"), colors.HexColor("#C7D2FE"), colors.HexColor("#4F46E5")),
            make_card(f"{replaced_count} Units", "Exchanged Claims", "Device replaced with inventory", colors.HexColor("#FFF7ED"), colors.HexColor("#FED7AA"), colors.HexColor("#D97706")),
            make_card(f"{refunded_count} Units", "Returned / Refunded", "Unfixed devices refunded/returned", colors.HexColor("#FEF2F2"), colors.HexColor("#FCA5A5"), colors.HexColor("#EF4444")),
            make_card(f"AED {total_cost:,.2f}", "Total Financial Impact", "Repair cost + refunds + balances", colors.HexColor("#ECFDF5"), colors.HexColor("#A7F3D0"), colors.HexColor("#059669"))
        ]]
        row1_table = Table(row1_data, colWidths=[66 * mm, 66 * mm, 66 * mm, 66 * mm])
        row1_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(row1_table)
        story.append(Spacer(1, 6 * mm))

        # Detailed Claims Table
        headers = [
            Paragraph("<b>Date</b>", th_style),
            Paragraph("<b>Customer Details</b>", th_left),
            Paragraph("<b>Service Item(s)</b>", th_left),
            Paragraph("<b>Details</b>", th_left),
            Paragraph("<b>Fault / Issue Note</b>", th_left),
            Paragraph("<b>Status</b>", th_style),
            Paragraph("<b>Resolution Details</b>", th_left),
            Paragraph("<b>Cost (AED)</b>", th_style),
        ]
        table_data = [headers]

        for idx, c in enumerate(filtered):
            claim_date_raw = c.get("claim_date") or ""
            cust_name = c.get("customer_name") or "—"
            cust_phone = c.get("phone_number") or ""
            cust_loc = c.get("location") or "N/A"
            cust_text = f"<b>{cust_name}</b><br/>Ph: {cust_phone}<br/>Loc: {cust_loc}"
            
            products_json_raw = c.get("products_json")
            p_list = []
            if products_json_raw:
                try:
                    p_list = json.loads(products_json_raw)
                except Exception:
                    pass
                    
            prod_lines = []
            if p_list:
                for p in p_list:
                    p_dta = p.get("dta") or ""
                    p_brand = p.get("brand") or ""
                    p_model = p.get("model") or ""
                    p_qty = p.get("qty") or 1
                    dta_txt = f"[{p_dta}] " if p_dta else ""
                    prod_lines.append(f"• {dta_txt}{p_brand} {p_model} (Qty: {p_qty})")
            else:
                dta_val = c.get("dta") or ""
                brand_val = c.get("brand") or ""
                model_val = c.get("model") or ""
                dta_txt = f"[{dta_val}] " if dta_val else ""
                prod_lines.append(f"• {dta_txt}{brand_val} {model_val} (Qty: 1)")
                
            prod_text = "<br/>".join(prod_lines)
            
            fulfillment = c.get("fulfillment_type") or "In-Store"
            w_status = c.get("warranty_status") or "Warranty"
            is_outside = int(c.get("is_outside") or 0) == 1
            detail_lines = [f"Fulfillment: {fulfillment}", f"Type: {w_status}"]
            if is_outside:
                detail_lines.append("<font color='#6B7280'><b>(Outside Product)</b></font>")
            detail_text = "<br/>".join(detail_lines)
            
            issue = c.get("issue_note") or ""
            
            status_val = c.get("status") or "Pending"
            status_badge = make_status_badge(status_val)
            
            res_lines = []
            cost_impact = 0.0
            if status_val != "Pending":
                action_date = c.get("action_date") or ""
                rep_note = c.get("repair_note") or ""
                rep_cost = float(c.get("repair_cost") or 0.0)
                exch_new_dta = c.get("exch_new_dta") or ""
                exch_new_brand = c.get("exch_new_brand") or ""
                exch_new_model = c.get("exch_new_model") or ""
                exch_balance = float(c.get("exch_balance") or 0.0)
                refund_amount = float(c.get("refund_amount") or 0.0)
                
                res_lines.append(f"<b>Date:</b> {action_date}")
                if exch_new_dta:
                    res_lines.append(f"<b>Exchanged For:</b> {exch_new_brand} {exch_new_model} ([{exch_new_dta}])")
                    if exch_balance != 0:
                        res_lines.append(f"<b>Exch Balance:</b> AED {exch_balance:,.2f}")
                if refund_amount > 0:
                    res_lines.append(f"<b>Refunded:</b> AED {refund_amount:,.2f}")
                if rep_note:
                    res_lines.append(f"<b>Notes:</b> {rep_note}")
                    
                cost_impact = rep_cost + exch_balance + refund_amount
            else:
                res_lines.append("<i>Awaiting diagnostics</i>")
                
            res_text = "<br/>".join(res_lines) if res_lines else "—"
            cost_text = f"AED {cost_impact:,.2f}" if cost_impact != 0 else "—"
            
            table_data.append([
                Paragraph(claim_date_raw, tdc),
                Paragraph(cust_text, td),
                Paragraph(prod_text, td),
                Paragraph(detail_text, td),
                Paragraph(issue, td),
                status_badge,
                Paragraph(res_text, td),
                Paragraph(cost_text, tdr),
            ])

        if len(filtered) == 0:
            table_data.append([
                Paragraph("<font color='#6c757d'>No warranty claims found matching the active filters.</font>", tdc),
                "", "", "", "", "", "", ""
            ])

        col_widths = [20 * mm, 40 * mm, 55 * mm, 30 * mm, 47 * mm, 22 * mm, 38 * mm, 15 * mm]
        table_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BRAND_COLOR),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("LINEBELOW", (0, 0), (-1, 0), 1.5, colors.HexColor("#312E81")),
            ("LINEBELOW", (0, 1), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ])
        
        if len(filtered) == 0:
            table_style.add("SPAN", (0, 1), (-1, 1))
            table_style.add("ALIGN", (0, 1), (-1, 1), "CENTER")
        else:
            for i in range(1, len(table_data)):
                if i % 2 == 1:
                    table_style.add("BACKGROUND", (0, i), (-1, i), ALT_ROW)
                else:
                    table_style.add("BACKGROUND", (0, i), (-1, i), WHITE)

        t = Table(table_data, colWidths=col_widths)
        t.setStyle(table_style)
        story.append(t)

        doc.build(story)
        buf.seek(0)

        filename = f"Warranty_Claims_Report_{date_cls.today().strftime('%Y-%m-%d')}.pdf"
        response = send_file(
            buf,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=filename
        )
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        return response

    except Exception as e:
        return jsonify({'error': str(e)}), 500


def parse_sticker_text(raw_text, pending_dta=""):
    import re
    lines = [l.strip() for l in raw_text.split("\n") if l.strip()]
    full_text = " ".join(lines)
    
    # 1. DTA Code
    dta = pending_dta
    dta_match = re.search(r'\bDTA[A-Z0-9]{3,10}\b', full_text, re.IGNORECASE)
    if dta_match:
        dta = dta_match.group(0).upper()
        
    # 2. Brand Detection
    known_brands = ["Lenovo", "Dell", "HP", "Asus", "Acer", "Apple", "MSI", "Samsung", "Toshiba", "Sony", "LG", "Huawei", "Microsoft", "Thinkpad"]
    brand = ""
    lower_text = full_text.lower()
    for b in known_brands:
        if b.lower() in lower_text:
            if b.lower() == "thinkpad":
                brand = "Lenovo"
            else:
                brand = b
            break
            
    # 3. Model Line Detection
    model_line = ""
    brand_keywords = [brand.lower()] if brand else []
    if brand == "Lenovo":
         brand_keywords.append("thinkpad")
         
    for l in lines:
        l_lower = l.lower()
        if any(kw in l_lower for kw in brand_keywords):
            model_line = l
            break
            
    if not model_line:
        clean_lines = [l for l in lines if not re.search(r'DTA[A-Z0-9]', l, re.IGNORECASE) and len(l) > 8]
        if clean_lines:
            model_line = sorted(clean_lines, key=len, reverse=True)[0]
            
    # Clean Model
    model = model_line
    extra_specs_from_model_line = []
    if "|" in model:
        parts = [p.strip() for p in model.split("|") if p.strip()]
        if len(parts) > 1:
            model = parts[0]
            extra_specs_from_model_line = parts[1:]

    if brand and model.lower().startswith(brand.lower()):
        model = model[len(brand):].strip()
    
    if brand == "Lenovo" and model.lower().startswith("lenovo"):
        model = model[6:].strip()
        
    model = re.sub(r'^[|/\-_\[\] ]+', '', model)
    model = re.sub(r'[|/\-_\[\] ]+$', '', model).strip()
    
    # 4. Specifications Extraction
    spec_keywords = ["intel", "amd", "core", "i3", "i5", "i7", "i9", "ryzen", "ram", "ssd", "gb", "tb", "hz", "ghz", "nvidia", "radeon", "graphics", "ddr", "celeron", "pentium", "generation", "gen", "pro"]
    spec_lines = []
    
    # Prepend any specs extracted from model line split
    for s in extra_specs_from_model_line:
        spec_lines.append(s)

    for l in lines:
        if l == model_line or (dta and dta in l.upper()):
            continue
        
        l_lower = l.lower()
        if "dithari" in l_lower or "ds193" in l_lower or l_lower.strip() == "d":
            continue
            
        if "|" in l:
            parts = [p.strip() for p in l.split("|") if p.strip()]
            for p in parts:
                p_lower = p.lower()
                is_spec = any(kw in p_lower for kw in spec_keywords) or \
                          re.search(r'\b\d+(?:gb|tb|g|t)\b', p_lower) or \
                          re.search(r'i[3579]-\d+', p_lower) or \
                          re.search(r'\b\d+\s*ssd\b', p_lower) or \
                          re.search(r'\b\d+\s*ram\b', p_lower)
                if is_spec:
                    cleaned = p
                    if brand and cleaned.lower().startswith(brand.lower()):
                        cleaned = cleaned[len(brand):].strip()
                    cleaned = re.sub(r'^[|/\-_\[\] ]+', '', cleaned)
                    cleaned = re.sub(r'[|/\-_\[\] ]+$', '', cleaned).strip()
                    if cleaned:
                        spec_lines.append(cleaned)
        else:
            is_spec = any(kw in l_lower for kw in spec_keywords) or \
                      re.search(r'\b\d+(?:gb|tb|g|t)\b', l_lower) or \
                      re.search(r'i[3579]-\d+', l_lower) or \
                      re.search(r'\b\d+\s*ssd\b', l_lower) or \
                      re.search(r'\b\d+\s*ram\b', l_lower)
            if is_spec:
                cleaned = l
                if brand and cleaned.lower().startswith(brand.lower()):
                    cleaned = cleaned[len(brand):].strip()
                cleaned = re.sub(r'^[|/\-_\[\] ]+', '', cleaned)
                cleaned = re.sub(r'[|/\-_\[\] ]+$', '', cleaned).strip()
                if cleaned:
                    spec_lines.append(cleaned)
                
    specs = " | ".join(spec_lines).strip()
    specs = re.sub(r'\s*\|\s*', ' | ', specs)
    
    if not brand and model:
        brand = model.split(" ")[0]
        
    return {
        "dta": dta,
        "brand": brand,
        "model": model,
        "specs": specs
    }


@app.route('/api/ocr/scan', methods=['POST'])
def api_ocr_scan():
    data = request.get_json()
    if not data or 'image' not in data:
        return jsonify({'error': 'Image data required'}), 400
        
    base64_image = data['image']
    pending_dta = data.get('pending_dta', '')
    api_key = data.get('api_key', 'K89034187088957') or 'K89034187088957'
    
    try:
        import urllib.request
        import urllib.parse
        import json
        
        post_data = urllib.parse.urlencode({
            "apikey": api_key,
            "language": "eng",
            "isOverlayRequired": "false",
            "OCREngine": "2",
            "base64Image": base64_image
        }).encode("utf-8")
        
        req = urllib.request.Request(
            "https://api.ocr.space/parse/image", 
            data=post_data, 
            headers={
                'User-Agent': 'Mozilla/5.0',
                'Content-Type': 'application/x-www-form-urlencoded'
            }
        )
        
        with urllib.request.urlopen(req, timeout=15) as r:
            res_data = r.read().decode('utf-8')
            res_json = json.loads(res_data)
            
            if res_json.get("IsErroredOnProcessing"):
                err_msg = res_json.get("ErrorMessage", ["OCR Service Error"])[0]
                return jsonify({'error': err_msg}), 400
                
            parsed_results = res_json.get("ParsedResults", [])
            if not parsed_results:
                return jsonify({'error': 'No text detected on the sticker'}), 400
                
            raw_text = parsed_results[0].get("ParsedText", "")
            if not raw_text.strip():
                return jsonify({'error': 'No text detected on the sticker'}), 400
                
            parsed_data = parse_sticker_text(raw_text, pending_dta)
            return jsonify(parsed_data)
            
    except Exception as e:
        return jsonify({'error': f"Failed to connect to OCR service: {str(e)}"}), 500


if __name__ == "__main__":
    import os
    import socket

    # Get current Wi-Fi IP
    wifi_ip = "127.0.0.1"
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        wifi_ip = s.getsockname()[0]
        s.close()
    except Exception:
        pass

    def run_tunnel():
        exe_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ngrok.exe")
        token = "3EWiuX0nXPImoEAVQR9OyDfSNdz_2ZgnNx1jeDPf32SAXwm5B"
        
        # 1. Download if missing
        if not os.path.exists(exe_path):
            print("\n[*] ngrok.exe is missing. Downloading (approx 20MB) for permanent mobile access...")
            try:
                import zipfile
                import io
                url = "https://bin.equinox.io/c/bNyj1mQVY4c/ngrok-v3-stable-windows-amd64.zip"
                req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
                zip_data = urllib.request.urlopen(req).read()
                with zipfile.ZipFile(io.BytesIO(zip_data)) as z:
                    z.extract("ngrok.exe", path=os.path.dirname(exe_path))
                print("[*] ngrok.exe downloaded and extracted successfully!")
            except Exception as e:
                print(f"[ERROR] Failed to download ngrok: {e}")
                return

        # 2. Configure authtoken
        try:
            subprocess.run([exe_path, "config", "add-authtoken", token], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception as e:
            print(f"[ERROR] Failed to configure ngrok authtoken: {e}")
            return

        # 3. Check for a custom permanent domain in ngrok_domain.txt
        domain_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ngrok_domain.txt")
        custom_domain = ""
        if os.path.exists(domain_file):
            try:
                with open(domain_file, "r") as f:
                    custom_domain = f.read().strip()
            except Exception:
                pass

        # 4. Start tunnel process
        cmd = [exe_path, "http", "5000"]
        if custom_domain:
            cmd.append(f"--domain={custom_domain}")
            
        try:
            proc = subprocess.Popen(
                cmd,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                text=True
            )
        except Exception as e:
            print(f"[ERROR] Failed to start ngrok process: {e}")
            return

        # 5. Retrieve public URL from local Ngrok API in background thread
        def capture_url():
            import json
            import time
            time.sleep(1.0)
            url = None
            for _ in range(15):
                try:
                    req = urllib.request.urlopen("http://127.0.0.1:4040/api/tunnels")
                    data = json.loads(req.read().decode())
                    tunnels = data.get("tunnels", [])
                    if tunnels:
                        for t in tunnels:
                            p_url = t.get("public_url", "")
                            if p_url.startswith("https://"):
                                url = p_url
                                break
                        if url:
                            break
                except Exception:
                    pass
                time.sleep(0.5)

            if url:
                print("\n" + "="*70)
                print(" SUCCESS: PERMANENT SECURE CLOUD TUNNEL ACTIVATED!")
                print(" Any mobile phone can now connect securely (camera works instantly!)")
                print(f" Mobile URL:  {url}")
                print("="*70 + "\n")
                
                # Write URL to tunnel_url.txt for easy copying
                try:
                    with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "tunnel_url.txt"), "w") as f:
                        f.write(url)
                except Exception:
                    pass

        import threading
        threading.Thread(target=capture_url, daemon=True).start()

    # Start the secure cloud tunnel in background
    import subprocess
    import urllib.request
    run_tunnel()

    print("\n" + "="*60)
    print("  BUYOLOGY BILLING SYSTEM — STARTED")
    print(f"  Laptop:  http://127.0.0.1:5000")
    print(f"  Phone / Other devices (Local Wi-Fi): http://{wifi_ip}:5000")
    print("="*60 + "\n")

    # Auto-open browser on laptop (plain HTTP locally is fast and has zero warnings)
    try:
        import webbrowser
        import threading
        threading.Timer(1.2, lambda: webbrowser.open("http://127.0.0.1:5000")).start()
    except Exception:
        pass

    # Run on fast plain HTTP locally
    app.run(debug=False, host="0.0.0.0", port=5000, threaded=True)
